import os, sys
import psycopg2
import psycopg2.extras
from datetime import date

STAGES = ['FIT UP', 'WELDING', 'BLASTING & PAINTING', 'SEND TO SITE']

WORKER_TYPES = [
    'Cutting Man', 'Supervisor', 'Foremen', 'Fitter', 'Welder',
    'Helper', 'Semi Skill', 'Material Coordinator', 'Material Handler',
]
SHIFT_LABELS = ['Regular', 'OT→6:30', 'OT→7:30', 'OT→10:00', 'Sun/PH']
SHIFT_KEYS   = ['regular', 'ot1', 'ot2', 'ot3', 'sun_ph']
SHIFT_HOURS  = {'regular': 7.5, 'ot1': 8.5, 'ot2': 9.5, 'ot3': 11.5, 'sun_ph': 7.5}


# ── PostgreSQL connection wrapper ──────────────────────────────────────────────

class _CurWrap:
    """Thin cursor wrapper: provides fetchone/fetchall/lastrowid over psycopg2 cursor."""
    def __init__(self, cur):
        self._cur = cur

    def fetchone(self):
        return self._cur.fetchone()  # RealDictRow (dict subclass) or None

    def fetchall(self):
        return self._cur.fetchall() or []  # list of RealDictRow

    @property
    def lastrowid(self):
        """Fetch the id returned by a RETURNING id clause."""
        row = self._cur.fetchone()
        return row['id'] if row else None


class _DBConn:
    """psycopg2 connection wrapper that mimics sqlite3 usage patterns.
    Converts ? placeholders to %s and uses RealDictCursor for dict-like row access.
    When created from the pool, close() returns the connection to the pool."""

    def __init__(self, conn, pool=None):
        self._conn = conn
        self._pool = pool

    def execute(self, sql, params=None):
        sql = sql.replace('?', '%s')
        cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params or [])
        return _CurWrap(cur)

    def commit(self):
        self._conn.commit()

    def close(self):
        if self._pool is not None:
            self._pool.putconn(self._conn)
        else:
            self._conn.close()


def _get_pool():
    """Return a module-level connection pool (created once per process)."""
    import streamlit as st
    import socket
    from urllib.parse import urlparse, unquote
    import psycopg2.pool

    url = st.secrets['database_url']
    p = urlparse(url)
    host = p.hostname
    try:
        host = socket.getaddrinfo(host, None, socket.AF_INET)[0][4][0]
    except Exception:
        pass

    return psycopg2.pool.ThreadedConnectionPool(
        1, 4,
        host=host,
        port=p.port or 5432,
        dbname=p.path.lstrip('/'),
        user=p.username,
        password=unquote(p.password or ''),
        sslmode='require',
        connect_timeout=10,
        keepalives=1,
        keepalives_idle=60,
        keepalives_interval=10,
        keepalives_count=5,
    )


_pool = None  # module-level singleton


def _conn():
    global _pool
    try:
        if _pool is None:
            _pool = _get_pool()
        conn = _pool.getconn()
        # Verify connection is alive; reset if stale
        try:
            conn.cursor().execute('SELECT 1')
        except Exception:
            try:
                _pool.putconn(conn, close=True)
            except Exception:
                pass
            conn = _pool.getconn()
        return _DBConn(conn, _pool)
    except Exception:
        # Pool failed — fall back to direct connection
        import streamlit as st
        import socket
        from urllib.parse import urlparse, unquote
        url = st.secrets['database_url']
        p = urlparse(url)
        host = p.hostname
        try:
            host = socket.getaddrinfo(host, None, socket.AF_INET)[0][4][0]
        except Exception:
            pass
        conn = psycopg2.connect(
            host=host, port=p.port or 5432, dbname=p.path.lstrip('/'),
            user=p.username, password=unquote(p.password or ''), sslmode='require',
        )
        return _DBConn(conn, pool=None)


# ── Schema initialisation ──────────────────────────────────────────────────────

def init():
    db = _conn()

    # Core tables
    db.execute("""
        CREATE TABLE IF NOT EXISTS assemblies (
            assembly_mark   TEXT PRIMARY KEY,
            total_weight_kg DOUBLE PRECISION DEFAULT 0,
            description     TEXT DEFAULT '',
            work_order      TEXT DEFAULT '001'
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS parts (
            id                SERIAL PRIMARY KEY,
            assembly_mark     TEXT NOT NULL,
            sub_assembly_mark TEXT DEFAULT '',
            part_mark         TEXT DEFAULT '',
            no                INTEGER DEFAULT 1,
            name              TEXT DEFAULT '',
            profile           TEXT DEFAULT '',
            kg_per_m          DOUBLE PRECISION DEFAULT 0,
            length_mm         DOUBLE PRECISION DEFAULT 0,
            total_weight_kg   DOUBLE PRECISION DEFAULT 0,
            profile2          TEXT DEFAULT '',
            grade             TEXT DEFAULT '',
            remark            TEXT DEFAULT '',
            FOREIGN KEY (assembly_mark) REFERENCES assemblies(assembly_mark)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS progress (
            id                SERIAL PRIMARY KEY,
            entry_date        TEXT NOT NULL,
            assembly_mark     TEXT NOT NULL,
            stage             TEXT NOT NULL,
            weight_kg         DOUBLE PRECISION DEFAULT 0,
            qty               INTEGER DEFAULT 0,
            inspector         TEXT DEFAULT '',
            remarks           TEXT DEFAULT '',
            created_at        TIMESTAMPTZ DEFAULT NOW(),
            sub_assembly_mark TEXT DEFAULT '',
            delivery_order_no TEXT DEFAULT ''
        )
    """)
    db.commit()

    # Idempotent column migrations
    for col_sql in [
        "ALTER TABLE progress   ADD COLUMN IF NOT EXISTS sub_assembly_mark TEXT DEFAULT ''",
        "ALTER TABLE progress   ADD COLUMN IF NOT EXISTS delivery_order_no TEXT DEFAULT ''",
        "ALTER TABLE parts      ADD COLUMN IF NOT EXISTS remark TEXT DEFAULT ''",
        "ALTER TABLE assemblies ADD COLUMN IF NOT EXISTS work_order TEXT DEFAULT '001'",
        "ALTER TABLE assemblies ADD COLUMN IF NOT EXISTS priority INTEGER DEFAULT 0",
        "ALTER TABLE parts      ADD COLUMN IF NOT EXISTS priority INTEGER",
        "ALTER TABLE progress   ADD COLUMN IF NOT EXISTS painting_done BOOLEAN DEFAULT FALSE",
    ]:
        db.execute(col_sql)
    db.commit()

    # Data migration: rename old DELIVERY stage
    db.execute("UPDATE progress SET stage='SEND TO SITE' WHERE stage='DELIVERY'")
    db.commit()

    # Indexes for frequently filtered columns
    for idx_sql in [
        "CREATE INDEX IF NOT EXISTS idx_progress_entry_date    ON progress(entry_date)",
        "CREATE INDEX IF NOT EXISTS idx_progress_assembly_mark ON progress(assembly_mark)",
        "CREATE INDEX IF NOT EXISTS idx_progress_stage         ON progress(stage)",
        "CREATE INDEX IF NOT EXISTS idx_parts_assembly_mark    ON parts(assembly_mark)",
        "CREATE INDEX IF NOT EXISTS idx_vi_assembly_mark       ON visual_inspection(assembly_mark)",
        "CREATE INDEX IF NOT EXISTS idx_vi_entry_date          ON visual_inspection(entry_date)",
    ]:
        db.execute(idx_sql)
    db.commit()

    # Sync assembly totals from parts (fixes stale values on startup)
    db.execute("""
        UPDATE assemblies
        SET total_weight_kg = (
            SELECT COALESCE(SUM(total_weight_kg), 0)
            FROM parts WHERE assembly_mark = assemblies.assembly_mark
        )
    """)
    db.commit()

    # Users table
    db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id            SERIAL PRIMARY KEY,
            username      TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role          TEXT DEFAULT 'user',
            active        INTEGER DEFAULT 1
        )
    """)
    db.commit()

    if db.execute("SELECT COUNT(*) AS cnt FROM users").fetchone()['cnt'] == 0:
        db.execute(
            "INSERT INTO users (username, password_hash, role) VALUES (?,?,?)",
            ('admin', _hash('admin123'), 'admin')
        )
        db.commit()

    # Manpower tables
    db.execute("""
        CREATE TABLE IF NOT EXISTS manpower (
            id          SERIAL PRIMARY KEY,
            entry_date  TEXT NOT NULL UNIQUE,
            regular     INTEGER DEFAULT 0,
            ot1         INTEGER DEFAULT 0,
            ot2         INTEGER DEFAULT 0,
            ot3         INTEGER DEFAULT 0,
            sun_ph      INTEGER DEFAULT 0,
            created_at  TIMESTAMPTZ DEFAULT NOW()
        )
    """)
    db.commit()

    db.execute("""
        CREATE TABLE IF NOT EXISTS manpower_detail (
            id          SERIAL PRIMARY KEY,
            entry_date  TEXT NOT NULL,
            worker_type TEXT NOT NULL,
            shift       TEXT NOT NULL,
            count       INTEGER DEFAULT 0,
            UNIQUE(entry_date, worker_type, shift)
        )
    """)
    db.commit()

    # Drawings table — files stored as BYTEA in the database
    db.execute("""
        CREATE TABLE IF NOT EXISTS drawings (
            id            SERIAL PRIMARY KEY,
            title         TEXT NOT NULL,
            original_name TEXT NOT NULL,
            filename      TEXT NOT NULL UNIQUE,
            assembly_mark TEXT DEFAULT '',
            uploaded_by   TEXT DEFAULT '',
            created_at    TIMESTAMPTZ DEFAULT NOW(),
            file_data     BYTEA
        )
    """)
    db.execute("ALTER TABLE drawings ADD COLUMN IF NOT EXISTS file_data BYTEA")
    db.execute("ALTER TABLE drawings ADD COLUMN IF NOT EXISTS rev_no TEXT DEFAULT ''")
    db.execute("ALTER TABLE drawings ADD COLUMN IF NOT EXISTS date_received TEXT DEFAULT ''")
    db.commit()

    # Project settings table
    db.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT DEFAULT ''
        )
    """)
    db.commit()

    db.close()
    init_raw_materials()
    init_visual_inspection()
    init_sessions()


def get_project_name():
    c = _conn()
    row = c.execute("SELECT value FROM settings WHERE key='project_name'").fetchone()
    c.close()
    return row['value'] if row else 'Fabrication Tracker'


def set_project_name(name):
    c = _conn()
    c.execute("""
        INSERT INTO settings (key, value) VALUES ('project_name', ?)
        ON CONFLICT(key) DO UPDATE SET value=EXCLUDED.value
    """, (name.strip(),))
    c.commit()
    c.close()


def _hash(password):
    import hashlib
    return hashlib.sha256(password.encode('utf-8')).hexdigest()


# ── Users ──────────────────────────────────────────────────────────────────────

def authenticate(username, password):
    c = _conn()
    row = c.execute(
        "SELECT id, username, role FROM users "
        "WHERE LOWER(username)=LOWER(?) AND password_hash=? AND active=1",
        (username.strip(), _hash(password))
    ).fetchone()
    c.close()
    return dict(row) if row else None


def get_users():
    c = _conn()
    rows = c.execute(
        "SELECT id, username, role, active FROM users ORDER BY username"
    ).fetchall()
    c.close()
    return [dict(r) for r in rows]


def add_user(username, password, role='user'):
    c = _conn()
    try:
        c.execute("INSERT INTO users (username, password_hash, role) VALUES (?,?,?)",
                  (username.strip(), _hash(password), role))
        c.commit()
        return True
    except Exception:
        return False
    finally:
        c.close()


def update_user_password(uid, new_password):
    c = _conn()
    c.execute("UPDATE users SET password_hash=? WHERE id=?", (_hash(new_password), uid))
    c.commit()
    c.close()


def update_user_role(uid, role):
    c = _conn()
    c.execute("UPDATE users SET role=? WHERE id=?", (role, uid))
    c.commit()
    c.close()


def toggle_user_active(uid):
    c = _conn()
    c.execute("UPDATE users SET active = 1 - active WHERE id=?", (uid,))
    c.commit()
    c.close()


def delete_user_entry(uid):
    c = _conn()
    c.execute("DELETE FROM users WHERE id=?", (uid,))
    c.commit()
    c.close()


# ── Raw Material Delivery ──────────────────────────────────────────────────────

def init_raw_materials():
    c = _conn()
    c.execute("""
        CREATE TABLE IF NOT EXISTS raw_materials (
            id            SERIAL PRIMARY KEY,
            received_date TEXT NOT NULL,
            do_no         TEXT DEFAULT '',
            description   TEXT DEFAULT '',
            grade         TEXT DEFAULT '',
            qty           DOUBLE PRECISION DEFAULT 0,
            total_kg      DOUBLE PRECISION DEFAULT 0,
            remark        TEXT DEFAULT '',
            created_at    TIMESTAMPTZ DEFAULT NOW()
        )
    """)
    c.execute("ALTER TABLE raw_materials ADD COLUMN IF NOT EXISTS do_no TEXT DEFAULT ''")
    c.execute("ALTER TABLE raw_materials ADD COLUMN IF NOT EXISTS total_kg DOUBLE PRECISION DEFAULT 0")
    c.commit()
    c.close()


def add_raw_material(received_date, do_no, description, grade, qty, total_kg=0, remark=''):
    c = _conn()
    cur = c.execute(
        "INSERT INTO raw_materials (received_date, do_no, description, grade, qty, total_kg, remark) "
        "VALUES (?,?,?,?,?,?,?) RETURNING id",
        (str(received_date), do_no.strip(), description.strip(), grade.strip(), qty, float(total_kg), remark.strip())
    )
    rid = cur.lastrowid
    c.commit()
    c.close()
    return rid


def get_raw_material_summary():
    """Return overall totals: total entries, total qty, total kg received."""
    c = _conn()
    row = c.execute(
        "SELECT COUNT(*) as entries, "
        "COALESCE(SUM(qty),0) as total_qty, "
        "COALESCE(SUM(total_kg),0) as total_kg "
        "FROM raw_materials"
    ).fetchone()
    c.close()
    return dict(row) if row else {'entries': 0, 'total_qty': 0, 'total_kg': 0}


def get_raw_materials(start=None, end=None):
    c = _conn()
    if start and end:
        rows = c.execute(
            "SELECT * FROM raw_materials WHERE received_date BETWEEN ? AND ? "
            "ORDER BY received_date DESC", (str(start), str(end))
        ).fetchall()
    else:
        rows = c.execute(
            "SELECT * FROM raw_materials ORDER BY received_date DESC"
        ).fetchall()
    c.close()
    return [dict(r) for r in rows]


def delete_raw_material(rid):
    c = _conn()
    c.execute("DELETE FROM raw_materials WHERE id=?", (rid,))
    c.commit()
    c.close()
    _reorder_raw_materials()


def _reorder_raw_materials():
    """Renumber raw_materials IDs sequentially after a deletion."""
    c = _conn()
    rows = c.execute(
        "SELECT received_date, do_no, description, grade, qty, total_kg, remark, created_at "
        "FROM raw_materials ORDER BY id"
    ).fetchall()
    c.execute("DELETE FROM raw_materials")
    c.execute("ALTER SEQUENCE raw_materials_id_seq RESTART WITH 1")
    for r in rows:
        c.execute(
            "INSERT INTO raw_materials "
            "(received_date, do_no, description, grade, qty, total_kg, remark, created_at) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (r['received_date'], r['do_no'], r['description'], r['grade'],
             r['qty'], r['total_kg'], r['remark'], r['created_at'])
        )
    c.commit()
    c.close()


def import_raw_materials_excel(file_source):
    """Import raw materials from Excel.
    file_source can be a file path (str) or bytes/BytesIO object.
    Expected columns (row 1 header): Received Date, D.O. Number, Description, Grade, Qty, Remark
    Returns (count, error_message). error_message is None on success.
    """
    try:
        import openpyxl
        from io import BytesIO as _BytesIO
        if isinstance(file_source, (bytes, bytearray)):
            file_source = _BytesIO(file_source)
        wb = openpyxl.load_workbook(file_source, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # find header row by looking for 'Description' or 'Received Date'
        header_row = next(
            (i for i, r in enumerate(rows)
             if r and any(str(v).strip().lower() in ('description', 'received date')
                          for v in r if v)),
            None
        )
        if header_row is None:
            return 0, "Header row not found. Ensure row 1 has: Received Date, D.O. Number, Description, Grade, Qty, Remark"
        headers = [str(h).strip() if h else '' for h in rows[header_row]]
        col = {h.lower(): i for i, h in enumerate(headers)}

        def _get(row, *names, default=''):
            for name in names:
                idx = col.get(name.lower())
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return row[idx]
            return default

        def _float(v):
            try: return float(v or 0)
            except: return 0.0

        c = _conn()
        count = 0
        for row in rows[header_row + 1:]:
            if not row or not any(v for v in row):
                continue
            desc = str(_get(row, 'description', default='')).strip()
            if not desc:
                continue
            recv     = str(_get(row, 'received date', 'received_date', default='')).strip()
            do_no    = str(_get(row, 'd.o. number', 'do number', 'do no', 'do_no', default='')).strip()
            grade    = str(_get(row, 'grade', default='')).strip()
            qty      = _float(_get(row, 'qty', 'quantity', default=0))
            total_kg = _float(_get(row, 'total kg', 'total_kg', 'total weight', default=0))
            remark   = str(_get(row, 'remark', 'remarks', default='')).strip()
            c.execute(
                "INSERT INTO raw_materials (received_date, do_no, description, grade, qty, total_kg, remark) "
                "VALUES (?,?,?,?,?,?,?)",
                (recv, do_no, desc, grade, qty, total_kg, remark)
            )
            count += 1
        c.commit()
        c.close()
        return count, None
    except Exception as e:
        return 0, str(e)


def replace_import_excel(file_source):
    """Clear all parts & assemblies (keeps progress), then reimport from Excel.
    file_source can be a file path (str) or bytes/BytesIO object."""
    c = _conn()
    # TRUNCATE is instant; DELETE on 4k-row tables hits Supabase statement timeout
    c.execute("TRUNCATE TABLE parts, assemblies")
    c.commit()
    c.close()
    return import_excel(file_source)


def import_excel(file_source):
    try:
        import openpyxl
        from io import BytesIO as _BytesIO
        if isinstance(file_source, (bytes, bytearray)):
            file_source = _BytesIO(file_source)
        wb = openpyxl.load_workbook(file_source, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header_row = next(
            (i for i, r in enumerate(rows)
             if r and any(str(v).strip() == 'Assembly Mark' for v in r if v is not None)),
            None
        )
        if header_row is None:
            return 0, 0, "Header row 'Assembly Mark' not found."
        headers = [str(h).strip() if h else '' for h in rows[header_row]]
        col = {h: i for i, h in enumerate(headers)}

        def _get(row, *names, default=None):
            for name in names:
                idx = col.get(name)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return row[idx]
            return default

        def _float(v):
            try: return float(v or 0)
            except: return 0.0

        def _int(v):
            try: return int(float(v or 1))
            except: return 1

        stage_cols = [
            ('FIT UP',              'FIT UP (kg)',               'FIT UP Date',                    None),
            ('WELDING',             'WELDING (kg)',              'WELDING Date',                   None),
            ('BLASTING & PAINTING', 'BLASTING & PAINTING (kg)', 'BLASTING & PAINTING Date',       'BLASTING & PAINTING D.O. No.'),
            ('SEND TO SITE',        'SEND TO SITE (kg)',         'SEND TO SITE Date',              'SEND TO SITE D.O. No.'),
        ]

        # ── Pass 1: parse entire Excel into memory lists ──────────────────────
        asm_order      = []        # insertion order preserved
        asm_set        = set()
        parts_rows     = []        # list of 12-tuples for bulk INSERT
        asm_weights    = {}        # asm -> total kg
        asm_work_orders = {}       # asm -> work_order
        asm_priorities  = {}       # asm -> priority
        progress_map   = {}        # (asm, sub, stage) -> (total_kg, date_str, do_no)

        for row in rows[header_row + 1:]:
            if not row or not _get(row, 'Assembly Mark'):
                continue

            asm    = str(_get(row, 'Assembly Mark', default='')).strip()
            sub    = str(_get(row, 'Sub Assembly', 'Sub-Assembly Mark', 'Sub Assembly Mark', default='') or '').strip()
            pm     = str(_get(row, 'Part Mark', default='') or '').strip()
            no     = _int(_get(row, 'No.', default=1))
            name   = str(_get(row, 'Name', 'NAME', default='') or '').strip()
            prof   = str(_get(row, 'Profile', default='') or '').strip()
            kgm    = _float(_get(row, 'kg/m', default=0))
            lmm    = _float(_get(row, 'Length (mm)', 'Length', default=0))
            tw     = _float(_get(row, 'Weight (kg)', 'Total weight', default=0))
            prof2  = str(_get(row, 'Profile 2', 'Profile2', default='') or '').strip()
            grade  = str(_get(row, 'Grade', default='') or '').strip()
            remark = str(_get(row, 'Remark', default='') or '').strip()
            wo     = str(_get(row, 'Work Order', 'Work_Order', 'WO', default='001') or '001').strip()
            _prio_raw = _get(row, 'Priority', default=None)
            prio = int(float(_prio_raw)) if _prio_raw is not None and str(_prio_raw).strip() != '' else None

            if asm not in asm_set:
                asm_order.append(asm)
                asm_set.add(asm)
                asm_work_orders[asm] = wo
                asm_priorities[asm]  = prio
            asm_weights[asm] = asm_weights.get(asm, 0) + tw
            parts_rows.append((asm, sub, pm, no, name, prof, kgm, lmm, tw, prof2, grade, remark, prio))

            for stage, kg_col, date_col, do_col in stage_cols:
                kg = _float(_get(row, kg_col, default=0))
                if kg > 0:
                    raw_date = _get(row, date_col, default=None)
                    if raw_date and hasattr(raw_date, 'strftime'):
                        date_str = raw_date.strftime('%Y-%m-%d')
                    elif raw_date:
                        date_str = str(raw_date).strip()[:10]
                    else:
                        date_str = str(date.today())
                    do_no = str(_get(row, do_col, default='') or '').strip() if do_col else ''
                    if (asm, sub, stage) in progress_map:
                        prev_kg, prev_date, prev_do = progress_map[(asm, sub, stage)]
                        progress_map[(asm, sub, stage)] = (prev_kg + kg, prev_date, prev_do or do_no)
                    else:
                        progress_map[(asm, sub, stage)] = (kg, date_str, do_no)

        if not parts_rows:
            return 0, 0, "No valid data rows found."

        # ── Pass 2: bulk DB writes (execute_values = one statement per batch) ─
        db  = _conn()
        raw = db._conn   # underlying psycopg2 connection
        cur = raw.cursor()

        # 1. Assemblies — all in one statement; include final weights and work_order
        psycopg2.extras.execute_values(
            cur,
            "INSERT INTO assemblies (assembly_mark, total_weight_kg, work_order, priority) VALUES %s "
            "ON CONFLICT(assembly_mark) DO UPDATE SET total_weight_kg = EXCLUDED.total_weight_kg, "
            "work_order = EXCLUDED.work_order, priority = EXCLUDED.priority",
            [(asm, asm_weights[asm], asm_work_orders.get(asm, '001'), asm_priorities.get(asm, None)) for asm in asm_order],
        )
        raw.commit()

        # 2. Parts — 500-row chunks (each chunk is one fast statement)
        CHUNK = 500
        for i in range(0, len(parts_rows), CHUNK):
            psycopg2.extras.execute_values(
                cur,
                "INSERT INTO parts "
                "(assembly_mark, sub_assembly_mark, part_mark, no, name, "
                "profile, kg_per_m, length_mm, total_weight_kg, profile2, grade, remark, priority) "
                "VALUES %s",
                parts_rows[i : i + CHUNK],
            )
            raw.commit()

        # 3. Progress — delete ALL entries for every imported assembly_mark, then insert
        # non-zero entries. Deleting by assembly_mark (not by stage+kg>0) ensures stale
        # entries from previous imports are fully replaced, even if kg is now 0.
        prog_rows  = [(ds, asm, sub, stg, kg, do_no)
                      for (asm, sub, stg), (kg, ds, do_no) in progress_map.items()]
        prog_count = 0
        if asm_order:
            psycopg2.extras.execute_values(
                cur,
                "DELETE FROM progress WHERE assembly_mark IN (SELECT v FROM (VALUES %s) AS t(v))",
                [(asm,) for asm in asm_order],
            )
            raw.commit()
        if prog_rows:
            psycopg2.extras.execute_values(
                cur,
                "INSERT INTO progress "
                "(entry_date, assembly_mark, sub_assembly_mark, stage, weight_kg, delivery_order_no) VALUES %s",
                prog_rows,
            )
            raw.commit()
            prog_count = len(prog_rows)

        cur.close()
        db.close()
        return len(parts_rows), prog_count, None
    except Exception as e:
        return 0, 0, str(e)


def get_work_orders():
    """Return sorted list of distinct work_order values from assemblies."""
    db = _conn()
    rows = db.execute(
        "SELECT DISTINCT work_order FROM assemblies WHERE work_order != '' ORDER BY work_order"
    ).fetchall()
    db.close()
    return [r['work_order'] for r in rows]


def get_marks_by_work_order(work_order=None):
    """Return assembly marks optionally filtered by work_order."""
    db = _conn()
    if work_order:
        rows = db.execute(
            "SELECT assembly_mark FROM assemblies WHERE work_order = ? ORDER BY assembly_mark",
            (work_order,)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT assembly_mark FROM assemblies ORDER BY assembly_mark"
        ).fetchall()
    db.close()
    return [r['assembly_mark'] for r in rows]


def add_assembly(mark, weight, desc='', work_order='001'):
    db = _conn()
    db.execute(
        "INSERT INTO assemblies (assembly_mark, total_weight_kg, description, work_order) VALUES (?, ?, ?, ?) "
        "ON CONFLICT(assembly_mark) DO NOTHING",
        (mark.strip().upper(), weight, desc, work_order.strip())
    )
    db.commit()
    db.close()


def add_part(asm, sub, pm, no, name, prof, kgm, lmm, tw, prof2, grade, remark=''):
    db = _conn()
    # ensure assembly exists
    db.execute(
        "INSERT INTO assemblies (assembly_mark, total_weight_kg) VALUES (?, 0) "
        "ON CONFLICT(assembly_mark) DO NOTHING",
        (asm,)
    )
    db.execute(
        "INSERT INTO parts (assembly_mark, sub_assembly_mark, part_mark, no, name, "
        "profile, kg_per_m, length_mm, total_weight_kg, profile2, grade, remark) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (asm, sub, pm, no, name, prof, kgm, lmm, tw, prof2, grade, remark)
    )
    # recalculate assembly total weight from sum of all its parts
    db.execute(
        "UPDATE assemblies SET total_weight_kg = "
        "(SELECT COALESCE(SUM(total_weight_kg), 0) FROM parts WHERE assembly_mark = ?) "
        "WHERE assembly_mark = ?",
        (asm, asm)
    )
    db.commit()
    db.close()


def update_part(pid, asm, sub, pm, no, name, prof, kgm, lmm, tw, prof2, grade, remark=''):
    db = _conn()
    old = db.execute("SELECT assembly_mark FROM parts WHERE id = ?", (pid,)).fetchone()
    old_asm = old['assembly_mark'] if old else None
    db.execute("""
        UPDATE parts SET assembly_mark=?, sub_assembly_mark=?, part_mark=?, no=?,
        name=?, profile=?, kg_per_m=?, length_mm=?, total_weight_kg=?, profile2=?, grade=?, remark=?
        WHERE id=?
    """, (asm, sub, pm, no, name, prof, kgm, lmm, tw, prof2, grade, remark, pid))
    # recalculate both old and new assembly weights if assembly changed
    for a in {asm, old_asm} - {None}:
        db.execute(
            "UPDATE assemblies SET total_weight_kg = "
            "(SELECT COALESCE(SUM(total_weight_kg),0) FROM parts WHERE assembly_mark=?) "
            "WHERE assembly_mark=?", (a, a)
        )
    db.commit()
    db.close()


def update_progress(pid, entry_date, mark, sub_mark, stage, weight, qty, remarks, do_no=''):
    db = _conn()
    db.execute("""
        UPDATE progress SET entry_date=?, assembly_mark=?, sub_assembly_mark=?, stage=?,
        weight_kg=?, qty=?, remarks=?, delivery_order_no=? WHERE id=?
    """, (str(entry_date), mark, sub_mark, stage, float(weight), int(qty), remarks, do_no, pid))
    db.commit()
    db.close()


def delete_part(part_id):
    db = _conn()
    row = db.execute("SELECT assembly_mark FROM parts WHERE id = ?", (part_id,)).fetchone()
    if row:
        asm = row['assembly_mark']
        db.execute("DELETE FROM parts WHERE id = ?", (part_id,))
        db.execute(
            "UPDATE assemblies SET total_weight_kg = "
            "(SELECT COALESCE(SUM(total_weight_kg), 0) FROM parts WHERE assembly_mark = ?) "
            "WHERE assembly_mark = ?",
            (asm, asm)
        )
        db.commit()
    db.close()


def get_marks():
    db = _conn()
    rows = db.execute("SELECT assembly_mark FROM assemblies ORDER BY assembly_mark").fetchall()
    db.close()
    return [r['assembly_mark'] for r in rows]


def get_assemblies():
    db = _conn()
    rows = db.execute("SELECT * FROM assemblies ORDER BY assembly_mark").fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_assembly_weight(mark):
    db = _conn()
    row = db.execute("SELECT total_weight_kg FROM assemblies WHERE assembly_mark = ?", (mark,)).fetchone()
    db.close()
    return row['total_weight_kg'] if row else 0


def progress_exists(entry_date, mark, sub_mark, stage):
    """Return True if a progress entry already exists for the given combination."""
    db = _conn()
    row = db.execute(
        "SELECT 1 AS exists FROM progress WHERE entry_date=? AND assembly_mark=? "
        "AND sub_assembly_mark=? AND stage=?",
        (str(entry_date), mark, sub_mark, stage)
    ).fetchone()
    db.close()
    return row is not None


def get_completed_stages(mark, sub_mark):
    """Return set of stages that have at least one progress entry for this assembly/sub-assembly."""
    db = _conn()
    rows = db.execute(
        "SELECT DISTINCT stage FROM progress "
        "WHERE assembly_mark=? AND sub_assembly_mark=?",
        (mark, sub_mark)
    ).fetchall()
    db.close()
    return {r['stage'] for r in rows}


def add_progress(entry_date, mark, sub_mark, stage, weight, qty, remarks, do_no=''):
    db = _conn()
    cur = db.execute(
        "INSERT INTO progress (entry_date, assembly_mark, sub_assembly_mark, stage, "
        "weight_kg, qty, remarks, delivery_order_no) VALUES (?, ?, ?, ?, ?, ?, ?, ?) RETURNING id",
        (str(entry_date), mark, sub_mark, stage, float(weight), int(qty), remarks, do_no)
    )
    rid = cur.lastrowid
    db.commit()
    db.close()
    return rid


def add_progress_bulk(entries):
    """Insert multiple progress rows in a single connection.
    entries: list of dicts with keys date, mark, sub, stage, weight, qty, remarks, do_no
    """
    if not entries:
        return
    conn = _conn()
    rows = [
        (str(e['date']), e['mark'], e['sub'], e['stage'],
         float(e['weight']), int(e['qty']), e['remarks'], e['do_no'])
        for e in entries
    ]
    psycopg2.extras.execute_values(
        conn._conn.cursor(),
        "INSERT INTO progress (entry_date, assembly_mark, sub_assembly_mark, stage, "
        "weight_kg, qty, remarks, delivery_order_no) VALUES %s",
        rows,
    )
    conn.commit()
    conn.close()


def delete_progress(rid):
    db = _conn()
    db.execute("DELETE FROM progress WHERE id = ?", (rid,))
    db.commit()
    db.close()


def clear_all_data():
    """Delete all records from progress, parts, and assemblies tables."""
    db = _conn()
    db.execute("DELETE FROM progress")
    db.execute("DELETE FROM parts")
    db.execute("DELETE FROM assemblies")
    db.commit()
    db.close()


def get_by_date(d):
    db = _conn()
    rows = db.execute(
        "SELECT p.*, a.total_weight_kg as asm_total FROM progress p "
        "JOIN assemblies a ON p.assembly_mark = a.assembly_mark "
        "WHERE p.entry_date = ? ORDER BY p.stage, p.assembly_mark",
        (str(d),)
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_by_range(start, end):
    db = _conn()
    rows = db.execute(
        "SELECT p.*, a.total_weight_kg as asm_total FROM progress p "
        "JOIN assemblies a ON p.assembly_mark = a.assembly_mark "
        "WHERE p.entry_date BETWEEN ? AND ? ORDER BY p.entry_date, p.stage, p.assembly_mark",
        (str(start), str(end))
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_cumulative():
    db = _conn()
    rows = db.execute("""
        SELECT a.assembly_mark, a.total_weight_kg,
            COALESCE(SUM(CASE WHEN p.stage='FIT UP'             THEN p.weight_kg END), 0) as fitup,
            COALESCE(SUM(CASE WHEN p.stage='WELDING'            THEN p.weight_kg END), 0) as welding,
            COALESCE(SUM(CASE WHEN p.stage='BLASTING & PAINTING' THEN p.weight_kg END), 0) as blasting,
            COALESCE(SUM(CASE WHEN p.stage='SEND TO SITE'       THEN p.weight_kg END), 0) as sendsite
        FROM assemblies a
        LEFT JOIN progress p ON a.assembly_mark = p.assembly_mark
        GROUP BY a.assembly_mark, a.total_weight_kg ORDER BY a.assembly_mark
    """).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_cumulative_by_sub(work_order=None):
    """Progress grouped by (assembly, sub-assembly).
    Total weight comes from the sub-assembly's parts weight.
    Assemblies with no sub-assemblies fall back to assembly-level totals.
    """
    db = _conn()
    params = []
    wo_filter1 = ''
    wo_filter2 = ''
    if work_order:
        wo_filter1 = 'AND a2.work_order = ?'
        wo_filter2 = 'AND a.work_order = ?'
        params = [work_order, work_order]
    rows = db.execute(f"""
        SELECT
            sp.assembly_mark,
            sp.sub_assembly_mark,
            sp.work_order,
            sp.priority,
            sp.sub_weight AS total_weight_kg,
            COALESCE(SUM(CASE WHEN p.stage='FIT UP'              THEN p.weight_kg END), 0) AS fitup,
            COALESCE(SUM(CASE WHEN p.stage='WELDING'             THEN p.weight_kg END), 0) AS welding,
            COALESCE(SUM(CASE WHEN p.stage='BLASTING & PAINTING' THEN p.weight_kg END), 0) AS blasting,
            COALESCE(SUM(CASE WHEN p.stage='SEND TO SITE'        THEN p.weight_kg END), 0) AS sendsite,
            MAX(CASE WHEN p.stage='BLASTING & PAINTING' THEN p.delivery_order_no END) AS blasting_do,
            MAX(CASE WHEN p.stage='SEND TO SITE'        THEN p.delivery_order_no END) AS sendsite_do
        FROM (
            SELECT pt.assembly_mark, pt.sub_assembly_mark,
                   a2.work_order, MAX(pt.priority) AS priority, SUM(pt.total_weight_kg) AS sub_weight
            FROM parts pt
            JOIN assemblies a2 ON pt.assembly_mark = a2.assembly_mark
            WHERE pt.sub_assembly_mark != ''
            {wo_filter1}
            GROUP BY pt.assembly_mark, pt.sub_assembly_mark, a2.work_order
        ) sp
        LEFT JOIN progress p
            ON sp.assembly_mark = p.assembly_mark
           AND sp.sub_assembly_mark = p.sub_assembly_mark
        GROUP BY sp.assembly_mark, sp.sub_assembly_mark, sp.work_order, sp.priority, sp.sub_weight

        UNION ALL

        SELECT
            a.assembly_mark,
            '' AS sub_assembly_mark,
            a.work_order,
            MAX(pt2.priority) AS priority,
            a.total_weight_kg,
            COALESCE(SUM(CASE WHEN p.stage='FIT UP'              THEN p.weight_kg END), 0) AS fitup,
            COALESCE(SUM(CASE WHEN p.stage='WELDING'             THEN p.weight_kg END), 0) AS welding,
            COALESCE(SUM(CASE WHEN p.stage='BLASTING & PAINTING' THEN p.weight_kg END), 0) AS blasting,
            COALESCE(SUM(CASE WHEN p.stage='SEND TO SITE'        THEN p.weight_kg END), 0) AS sendsite,
            MAX(CASE WHEN p.stage='BLASTING & PAINTING' THEN p.delivery_order_no END) AS blasting_do,
            MAX(CASE WHEN p.stage='SEND TO SITE'        THEN p.delivery_order_no END) AS sendsite_do
        FROM assemblies a
        LEFT JOIN parts pt2 ON a.assembly_mark = pt2.assembly_mark
        LEFT JOIN progress p ON a.assembly_mark = p.assembly_mark
        WHERE a.assembly_mark NOT IN (
            SELECT DISTINCT assembly_mark FROM parts WHERE sub_assembly_mark != ''
        )
        {wo_filter2}
        GROUP BY a.assembly_mark, a.work_order, a.total_weight_kg
        ORDER BY 1, 2
    """, params).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_daily_production():
    """Return kg produced per entry_date per stage — for trend and S-curve charts."""
    c = _conn()
    rows = c.execute("""
        SELECT entry_date, stage, COALESCE(SUM(weight_kg), 0) AS kg
        FROM progress
        GROUP BY entry_date, stage
        ORDER BY entry_date, stage
    """).fetchall()
    c.close()
    return [dict(r) for r in rows]


def get_daily_manhours():
    """Return total manhours per entry_date from manpower_detail."""
    c = _conn()
    rows = c.execute("""
        SELECT entry_date, shift, COALESCE(SUM(count), 0) AS headcount
        FROM manpower_detail
        GROUP BY entry_date, shift
        ORDER BY entry_date
    """).fetchall()
    c.close()
    result = {}
    for r in rows:
        d = r['entry_date']
        result[d] = result.get(d, 0) + r['headcount'] * SHIFT_HOURS.get(r['shift'], 0)
    return [{'entry_date': k, 'manhours': v} for k, v in sorted(result.items())]


def get_stage_daily_stats():
    """Return total_kg and unique day count per stage — single aggregate query."""
    c = _conn()
    rows = c.execute("""
        SELECT p.stage,
               COALESCE(SUM(p.weight_kg), 0)        AS total_kg,
               COUNT(DISTINCT p.entry_date)          AS days
        FROM progress p
        JOIN assemblies a ON p.assembly_mark = a.assembly_mark
        GROUP BY p.stage
    """).fetchall()
    c.close()
    result = {}
    for r in rows:
        days  = r['days']     or 0
        total = r['total_kg'] or 0
        result[r['stage']] = {
            'total_kg':    total,
            'days':        days,
            'avg_per_day': total / days if days else 0,
        }
    return result


def get_summary(as_of_date=None):
    db = _conn()
    total = db.execute(
        "SELECT COALESCE(SUM(total_weight_kg),0) AS total FROM assemblies"
    ).fetchone()['total']
    if as_of_date:
        rows = db.execute(
            "SELECT p.stage, COALESCE(SUM(p.weight_kg),0) as done "
            "FROM progress p JOIN assemblies a ON p.assembly_mark = a.assembly_mark "
            "WHERE p.entry_date <= ? "
            "GROUP BY p.stage",
            (str(as_of_date),)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT p.stage, COALESCE(SUM(p.weight_kg),0) as done "
            "FROM progress p JOIN assemblies a ON p.assembly_mark = a.assembly_mark "
            "GROUP BY p.stage"
        ).fetchall()
    db.close()
    result = {'total': total}
    for r in rows:
        result[r['stage']] = r['done']
    return result


def get_all_daily_stage_totals():
    """Return list of {entry_date, stage, kg} for every date+stage that has progress.
    Used by the Report tab to compute both cumulative and daily totals in Python,
    avoiding a new DB round-trip each time the user changes the selected date."""
    db = _conn()
    rows = db.execute(
        "SELECT p.entry_date, p.stage, COALESCE(SUM(p.weight_kg),0) AS kg "
        "FROM progress p JOIN assemblies a ON p.assembly_mark = a.assembly_mark "
        "GROUP BY p.entry_date, p.stage"
    ).fetchall()
    db.close()
    return [{'entry_date': str(r['entry_date']), 'stage': r['stage'], 'kg': float(r['kg'] or 0)} for r in rows]


def get_on_hold_weight():
    """Total weight_kg of parts whose remark contains 'on hold' (case-insensitive)."""
    db = _conn()
    row = db.execute(
        "SELECT COALESCE(SUM(total_weight_kg), 0) AS kg "
        "FROM parts WHERE UPPER(remark) LIKE ? OR UPPER(remark) LIKE ?",
        ('%ON HOLD%', '%ON-HOLD%')
    ).fetchone()
    db.close()
    return row['kg'] if row else 0


def get_sub_assemblies(assembly_mark):
    """Return distinct sub-assembly marks for a given assembly."""
    db = _conn()
    rows = db.execute(
        "SELECT DISTINCT sub_assembly_mark FROM parts "
        "WHERE assembly_mark = ? AND sub_assembly_mark != '' "
        "ORDER BY sub_assembly_mark",
        (assembly_mark,)
    ).fetchall()
    db.close()
    return [r['sub_assembly_mark'] for r in rows]


def get_deliveries():
    db = _conn()
    rows = db.execute(
        "SELECT p.id, p.entry_date, a.work_order, p.assembly_mark, p.sub_assembly_mark, p.stage, "
        "p.delivery_order_no, p.weight_kg, p.qty, p.remarks, "
        "COALESCE(p.painting_done, FALSE) AS painting_done "
        "FROM progress p "
        "JOIN assemblies a ON p.assembly_mark = a.assembly_mark "
        "WHERE p.stage IN ('BLASTING & PAINTING','SEND TO SITE') "
        "ORDER BY p.entry_date DESC, p.assembly_mark"
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def set_painting_done(progress_id, done: bool):
    db = _conn()
    db.execute("UPDATE progress SET painting_done = ? WHERE id = ?", (done, progress_id))
    db.commit()
    db.close()


def get_painting_done_kg(up_to_date: str = None):
    """Total weight_kg of B&P entries marked painting_done, optionally up to a date."""
    db = _conn()
    if up_to_date:
        row = db.execute(
            "SELECT COALESCE(SUM(weight_kg), 0) AS kg FROM progress "
            "WHERE stage = 'BLASTING & PAINTING' AND painting_done = TRUE "
            "AND entry_date <= ?", (up_to_date,)
        ).fetchone()
    else:
        row = db.execute(
            "SELECT COALESCE(SUM(weight_kg), 0) AS kg FROM progress "
            "WHERE stage = 'BLASTING & PAINTING' AND painting_done = TRUE"
        ).fetchone()
    db.close()
    return float(row['kg']) if row else 0.0


def set_painting_done_by_do(do_no: str, done: bool):
    db = _conn()
    db.execute(
        "UPDATE progress SET painting_done = ? "
        "WHERE stage = 'BLASTING & PAINTING' AND delivery_order_no = ?",
        (done, do_no)
    )
    db.commit()
    db.close()


def get_parts(assembly_mark=None):
    db = _conn()
    if assembly_mark:
        rows = db.execute(
            "SELECT * FROM parts WHERE assembly_mark = ? "
            "ORDER BY assembly_mark, part_mark", (assembly_mark,)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM parts ORDER BY assembly_mark, part_mark"
        ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_master_export():
    """Parts table joined with cumulative progress per (assembly, sub-assembly)."""
    db = _conn()
    rows = db.execute("""
        SELECT
            p.priority             AS "Priority",
            a.work_order           AS "Work Order",
            p.assembly_mark        AS "Assembly Mark",
            p.sub_assembly_mark    AS "Sub Assembly",
            p.part_mark            AS "Part Mark",
            p.no                   AS "No.",
            p.name                 AS "Name",
            p.profile              AS "Profile",
            p.kg_per_m             AS "kg/m",
            p.length_mm            AS "Length (mm)",
            p.total_weight_kg      AS "Weight (kg)",
            p.profile2             AS "Profile 2",
            p.grade                AS "Grade",
            p.remark               AS "Remark",
            CASE WHEN pr.fitup_done    = 1 THEN p.total_weight_kg ELSE 0 END AS "FIT UP (kg)",
            pr.fitup_dates                                                   AS "FIT UP Date",
            CASE WHEN pr.welding_done  = 1 THEN p.total_weight_kg ELSE 0 END AS "WELDING (kg)",
            pr.welding_dates                                                 AS "WELDING Date",
            CASE WHEN pr.blasting_done = 1 THEN p.total_weight_kg ELSE 0 END AS "BLASTING & PAINTING (kg)",
            pr.blasting_dates                                                AS "BLASTING & PAINTING Date",
            pr.blasting_do                                                   AS "BLASTING & PAINTING D.O. No.",
            CASE WHEN pr.sendsite_done = 1 THEN p.total_weight_kg ELSE 0 END AS "SEND TO SITE (kg)",
            pr.sendsite_dates                                                AS "SEND TO SITE Date",
            pr.sendsite_do                                                   AS "SEND TO SITE D.O. No."
        FROM parts p
        JOIN assemblies a ON p.assembly_mark = a.assembly_mark
        LEFT JOIN (
            SELECT
                assembly_mark,
                sub_assembly_mark,
                MAX(CASE WHEN stage='FIT UP'              THEN 1 ELSE 0 END) AS fitup_done,
                MAX(CASE WHEN stage='WELDING'             THEN 1 ELSE 0 END) AS welding_done,
                MAX(CASE WHEN stage='BLASTING & PAINTING' THEN 1 ELSE 0 END) AS blasting_done,
                MAX(CASE WHEN stage='SEND TO SITE'        THEN 1 ELSE 0 END) AS sendsite_done,
                STRING_AGG(CASE WHEN stage='FIT UP'              THEN entry_date END, ',') AS fitup_dates,
                STRING_AGG(CASE WHEN stage='WELDING'             THEN entry_date END, ',') AS welding_dates,
                STRING_AGG(CASE WHEN stage='BLASTING & PAINTING' THEN entry_date END, ',') AS blasting_dates,
                STRING_AGG(CASE WHEN stage='SEND TO SITE'        THEN entry_date END, ',') AS sendsite_dates,
                MAX(CASE WHEN stage='BLASTING & PAINTING' THEN delivery_order_no END)      AS blasting_do,
                MAX(CASE WHEN stage='SEND TO SITE'        THEN delivery_order_no END)      AS sendsite_do
            FROM progress
            GROUP BY assembly_mark, sub_assembly_mark
        ) pr ON p.assembly_mark = pr.assembly_mark
             AND p.sub_assembly_mark = pr.sub_assembly_mark
        ORDER BY p.assembly_mark, p.sub_assembly_mark, p.part_mark
    """).fetchall()
    db.close()
    return [dict(r) for r in rows]


def get_parts_summary(assembly_mark):
    """Return part count and total weight for an assembly."""
    db = _conn()
    row = db.execute(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(total_weight_kg),0) as total "
        "FROM parts WHERE assembly_mark = ?", (assembly_mark,)
    ).fetchone()
    db.close()
    return dict(row) if row else {'cnt': 0, 'total': 0}


def search_parts(keyword='', assembly_mark=None):
    """Search parts by keyword across all text columns."""
    db = _conn()
    kw = f'%{keyword}%'
    if assembly_mark:
        rows = db.execute("""
            SELECT * FROM parts
            WHERE assembly_mark = ?
              AND (assembly_mark ILIKE ? OR sub_assembly_mark ILIKE ? OR part_mark ILIKE ?
                   OR name ILIKE ? OR profile ILIKE ? OR profile2 ILIKE ? OR grade ILIKE ?)
            ORDER BY assembly_mark, part_mark
        """, (assembly_mark, kw, kw, kw, kw, kw, kw, kw)).fetchall()
    else:
        rows = db.execute("""
            SELECT * FROM parts
            WHERE (assembly_mark ILIKE ? OR sub_assembly_mark ILIKE ? OR part_mark ILIKE ?
                   OR name ILIKE ? OR profile ILIKE ? OR profile2 ILIKE ? OR grade ILIKE ?)
            ORDER BY assembly_mark, part_mark
        """, (kw, kw, kw, kw, kw, kw, kw)).fetchall()
    db.close()
    return [dict(r) for r in rows]


def search_progress(keyword='', stage=None, assembly_mark=None, start=None, end=None, work_order=None):
    """Search progress entries by keyword, stage, assembly, date range, and/or work_order."""
    db = _conn()
    kw = f'%{keyword}%'
    conditions = ["(p.assembly_mark ILIKE ? OR p.remarks ILIKE ?)"]
    params = [kw, kw]
    if stage:
        conditions.append("p.stage = ?")
        params.append(stage)
    if assembly_mark:
        conditions.append("p.assembly_mark = ?")
        params.append(assembly_mark)
    if start:
        conditions.append("p.entry_date >= ?")
        params.append(str(start))
    if end:
        conditions.append("p.entry_date <= ?")
        params.append(str(end))
    if work_order:
        conditions.append("a.work_order = ?")
        params.append(work_order)
    where = " AND ".join(conditions)
    rows = db.execute(f"""
        SELECT p.*, a.total_weight_kg as asm_total, a.work_order
        FROM progress p
        JOIN assemblies a ON p.assembly_mark = a.assembly_mark
        WHERE {where}
        ORDER BY p.entry_date DESC, p.stage, p.assembly_mark
    """, params).fetchall()
    db.close()
    return [dict(r) for r in rows]


def export_csv(rows, path):
    import csv
    if not rows:
        return
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        w.writeheader()
        w.writerows(rows)


def export_excel(rows, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    if not rows:
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Progress'
    headers = list(rows[0].keys())
    # header row styling
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h.replace('_', ' ').title())
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
    # data rows with alternating fill
    alt_fill = PatternFill('solid', fgColor='EEF2FF')
    for row_i, r in enumerate(rows, 2):
        for col, h in enumerate(headers, 1):
            ws.cell(row=row_i, column=col, value=r.get(h, ''))
        if row_i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_i, column=col).fill = alt_fill
    # auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    wb.save(path)


# ── Visual Inspection ─────────────────────────────────────────────────────────

def init_visual_inspection():
    c = _conn()
    c.execute("""
        CREATE TABLE IF NOT EXISTS visual_inspection (
            id                SERIAL PRIMARY KEY,
            entry_date        TEXT NOT NULL,
            assembly_mark     TEXT DEFAULT '',
            sub_assembly_mark TEXT DEFAULT '',
            weight_kg         DOUBLE PRECISION DEFAULT 0,
            qty               INTEGER DEFAULT 1,
            remarks           TEXT DEFAULT '',
            created_at        TIMESTAMPTZ DEFAULT NOW()
        )
    """)
    c.commit()
    c.close()


def visual_inspection_passed(mark, sub_mark):
    """Return True if at least one visual inspection record exists for this assembly/sub-assembly."""
    c = _conn()
    row = c.execute(
        "SELECT 1 FROM visual_inspection WHERE assembly_mark=? AND sub_assembly_mark=?",
        (mark.strip().upper(), sub_mark.strip().upper())
    ).fetchone()
    c.close()
    return row is not None


def visual_inspection_exists(entry_date, mark, sub_mark):
    """Return True if a record already exists for this date/assembly/sub-assembly."""
    c = _conn()
    row = c.execute(
        "SELECT 1 FROM visual_inspection "
        "WHERE entry_date=? AND assembly_mark=? AND sub_assembly_mark=?",
        (str(entry_date), mark.strip().upper(), sub_mark.strip().upper())
    ).fetchone()
    c.close()
    return row is not None


def add_visual_inspection(entry_date, mark, sub_mark, weight_kg, qty, remarks=''):
    c = _conn()
    cur = c.execute(
        "INSERT INTO visual_inspection (entry_date, assembly_mark, sub_assembly_mark, "
        "weight_kg, qty, remarks) VALUES (?,?,?,?,?,?) RETURNING id",
        (str(entry_date), mark.strip().upper(), sub_mark.strip().upper(),
         float(weight_kg), int(qty), remarks.strip())
    )
    rid = cur.lastrowid
    c.commit()
    c.close()
    return rid


def bulk_add_visual_inspection(entry_date, records):
    """Insert multiple VI records in one connection.
    records: list of dicts with keys mark, sub, weight_kg, qty, remarks.
    Skips duplicates (same date/mark/sub already exists).
    Returns count of rows inserted.
    """
    if not records:
        return 0
    c = _conn()
    count = 0
    for r in records:
        exists = c.execute(
            "SELECT 1 FROM visual_inspection "
            "WHERE entry_date=? AND assembly_mark=? AND sub_assembly_mark=?",
            (str(entry_date), r['mark'].upper(), r['sub'].upper())
        ).fetchone()
        if not exists:
            c.execute(
                "INSERT INTO visual_inspection "
                "(entry_date, assembly_mark, sub_assembly_mark, weight_kg, qty, remarks) "
                "VALUES (?,?,?,?,?,?)",
                (str(entry_date), r['mark'].upper(), r['sub'].upper(),
                 float(r['weight_kg']), int(r.get('qty', 1)), r.get('remarks', ''))
            )
            count += 1
    c.commit()
    c.close()
    return count


def get_visual_inspections(start=None, end=None):
    c = _conn()
    if start and end:
        rows = c.execute(
            "SELECT * FROM visual_inspection WHERE entry_date BETWEEN ? AND ? "
            "ORDER BY entry_date DESC, id DESC",
            (str(start), str(end))
        ).fetchall()
    else:
        rows = c.execute(
            "SELECT * FROM visual_inspection ORDER BY entry_date DESC, id DESC"
        ).fetchall()
    c.close()
    return [dict(r) for r in rows]


def get_visual_inspection_summary():
    c = _conn()
    row = c.execute(
        "SELECT COUNT(*) as entries, COALESCE(SUM(weight_kg),0) as total_kg "
        "FROM visual_inspection"
    ).fetchone()
    c.close()
    return dict(row) if row else {'entries': 0, 'total_kg': 0}


def get_missing_visual_inspections():
    """Return sub-assemblies that have WELDING recorded but no VI record yet."""
    c = _conn()
    rows = c.execute("""
        SELECT p.assembly_mark, p.sub_assembly_mark,
               COALESCE(SUM(p.weight_kg), 0) AS welding_kg
        FROM progress p
        WHERE p.stage = 'WELDING'
          AND NOT EXISTS (
              SELECT 1 FROM visual_inspection vi
              WHERE vi.assembly_mark     = p.assembly_mark
                AND vi.sub_assembly_mark = p.sub_assembly_mark
          )
        GROUP BY p.assembly_mark, p.sub_assembly_mark
        ORDER BY p.assembly_mark, p.sub_assembly_mark
    """).fetchall()
    c.close()
    return [dict(r) for r in rows]


def delete_visual_inspection(rid):
    c = _conn()
    c.execute("DELETE FROM visual_inspection WHERE id=?", (rid,))
    c.commit()
    c.close()


def import_visual_inspection_excel(file_source):
    """Import visual inspection records from Excel.
    Expected columns: Date, Assembly Mark, Sub Assembly Mark, Weight (kg), Qty, Remarks
    Skips duplicates (same date + assembly + sub-assembly).
    Returns (inserted, skipped, error_message).
    """
    try:
        import openpyxl
        from io import BytesIO as _BytesIO
        if isinstance(file_source, (bytes, bytearray)):
            file_source = _BytesIO(file_source)
        wb = openpyxl.load_workbook(file_source, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        header_row = next(
            (i for i, r in enumerate(rows)
             if r and any(str(v).strip().lower() in ('date', 'assembly mark')
                          for v in r if v is not None)),
            None
        )
        if header_row is None:
            return 0, 0, "Header row not found. Ensure columns: Date, Assembly Mark, Sub Assembly Mark, Weight (kg), Qty, Remarks"

        headers = [str(h).strip().lower() if h else '' for h in rows[header_row]]
        col = {h: i for i, h in enumerate(headers)}

        def _get(row, *names, default=None):
            for name in names:
                idx = col.get(name.lower())
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return row[idx]
            return default

        def _float(v):
            try: return float(v)
            except: return 0.0

        def _int(v):
            try: return int(v)
            except: return 1

        c = _conn()
        inserted = 0
        skipped  = 0
        for row in rows[header_row + 1:]:
            if not row or not any(v for v in row):
                continue
            raw_date = _get(row, 'date', default=None)
            if raw_date and hasattr(raw_date, 'strftime'):
                entry_date = raw_date.strftime('%Y-%m-%d')
            elif raw_date:
                entry_date = str(raw_date).strip()
            else:
                continue

            mark = str(_get(row, 'assembly mark', default='') or '').strip().upper()
            if not mark:
                continue
            sub  = str(_get(row, 'sub assembly mark', 'sub assembly', 'sub-assembly mark', default='') or '').strip().upper()
            wt   = _float(_get(row, 'weight (kg)', 'weight', 'kg', default=0))
            qty  = _int(_get(row, 'qty', 'quantity', default=1))
            rmk  = str(_get(row, 'remarks', 'remark', default='') or '').strip()

            # skip duplicates
            dup = c.execute(
                "SELECT 1 FROM visual_inspection WHERE entry_date=? AND assembly_mark=? AND sub_assembly_mark=?",
                (entry_date, mark, sub)
            ).fetchone()
            if dup:
                skipped += 1
                continue

            c.execute(
                "INSERT INTO visual_inspection (entry_date, assembly_mark, sub_assembly_mark, weight_kg, qty, remarks) "
                "VALUES (?,?,?,?,?,?)",
                (entry_date, mark, sub, wt, qty, rmk)
            )
            inserted += 1

        c.commit()
        c.close()
        return inserted, skipped, None
    except Exception as e:
        return 0, 0, str(e)


# ── Session / Online Tracking ──────────────────────────────────────────────────

def init_sessions():
    c = _conn()
    c.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            id         SERIAL PRIMARY KEY,
            username   TEXT NOT NULL,
            role       TEXT DEFAULT '',
            login_time TEXT NOT NULL,
            last_seen  TEXT NOT NULL,
            active     INTEGER DEFAULT 1
        )
    """)
    c.commit()
    c.close()


def _now_gmt8():
    from datetime import datetime as _dt, timedelta as _td
    return (_dt.utcnow() + _td(hours=8)).strftime('%Y-%m-%d %H:%M:%S')


def create_session(username, role=''):
    now = _now_gmt8()
    c = _conn()
    cur = c.execute(
        "INSERT INTO sessions (username, role, login_time, last_seen) VALUES (?,?,?,?) RETURNING id",
        (username, role, now, now)
    )
    sid = cur.lastrowid
    c.commit()
    c.close()
    return sid


def update_session_heartbeat(session_id):
    now = _now_gmt8()
    c = _conn()
    c.execute("UPDATE sessions SET last_seen=? WHERE id=?", (now, session_id))
    c.commit()
    c.close()


def end_session(session_id):
    c = _conn()
    c.execute("UPDATE sessions SET active=0 WHERE id=?", (session_id,))
    c.commit()
    c.close()


def get_active_sessions(minutes=10):
    """Users active within the last N minutes (GMT+8)."""
    from datetime import datetime as _dt, timedelta as _td
    threshold = (_dt.utcnow() + _td(hours=8) - _td(minutes=minutes)).strftime('%Y-%m-%d %H:%M:%S')
    c = _conn()
    rows = c.execute(
        "SELECT username, role, login_time, last_seen FROM sessions "
        "WHERE active=1 AND last_seen >= ? "
        "ORDER BY last_seen DESC",
        (threshold,)
    ).fetchall()
    c.close()
    return [dict(r) for r in rows]


def get_login_history(limit=100):
    """Recent login sessions, newest first."""
    c = _conn()
    rows = c.execute(
        "SELECT username, role, login_time, last_seen, active FROM sessions "
        "ORDER BY login_time DESC LIMIT ?",
        (limit,)
    ).fetchall()
    c.close()
    return [dict(r) for r in rows]


# ── Manpower ───────────────────────────────────────────────────────────────────

def save_manpower(entry_date, regular, ot1, ot2, ot3, sun_ph,
                  cutting_man=0, supervisor=0, foremen=0, fitter=0, helper=0, semi_skill=0,
                  material_coordinator=0, material_handler=0):
    c = _conn()
    c.execute("""
        INSERT INTO manpower
            (entry_date, regular, ot1, ot2, ot3, sun_ph,
             cutting_man, supervisor, foremen, fitter, helper, semi_skill,
             material_coordinator, material_handler)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(entry_date) DO UPDATE SET
            regular=EXCLUDED.regular, ot1=EXCLUDED.ot1, ot2=EXCLUDED.ot2,
            ot3=EXCLUDED.ot3, sun_ph=EXCLUDED.sun_ph,
            cutting_man=EXCLUDED.cutting_man, supervisor=EXCLUDED.supervisor,
            foremen=EXCLUDED.foremen, fitter=EXCLUDED.fitter,
            helper=EXCLUDED.helper, semi_skill=EXCLUDED.semi_skill,
            material_coordinator=EXCLUDED.material_coordinator,
            material_handler=EXCLUDED.material_handler
    """, (str(entry_date),
          int(regular), int(ot1), int(ot2), int(ot3), int(sun_ph),
          int(cutting_man), int(supervisor), int(foremen),
          int(fitter), int(helper), int(semi_skill),
          int(material_coordinator), int(material_handler)))
    c.commit()
    c.close()


def get_manpower(entry_date):
    c = _conn()
    row = c.execute("SELECT * FROM manpower WHERE entry_date=?", (str(entry_date),)).fetchone()
    c.close()
    return dict(row) if row else None


def save_manpower_grid(entry_date, grid):
    """Save a full grid {worker_type: {shift_key: count}} for a date."""
    c = _conn()
    c.execute("DELETE FROM manpower_detail WHERE entry_date=?", (str(entry_date),))
    for wtype, shifts in grid.items():
        for shift_key, count in shifts.items():
            if int(count) > 0:
                c.execute("""
                    INSERT INTO manpower_detail (entry_date, worker_type, shift, count)
                    VALUES (?,?,?,?)
                """, (str(entry_date), wtype, shift_key, int(count)))
    c.commit()
    c.close()


def get_manpower_grid(entry_date):
    """Return {worker_type: {shift_key: count}} for a date."""
    c = _conn()
    rows = c.execute(
        "SELECT worker_type, shift, count FROM manpower_detail WHERE entry_date=?",
        (str(entry_date),)
    ).fetchall()
    c.close()
    grid = {}
    for r in rows:
        grid.setdefault(r['worker_type'], {})[r['shift']] = r['count']
    return grid


def get_manhour_summary():
    """Total manhours, total days logged, and average manhours per day."""
    c = _conn()
    rows = c.execute(
        "SELECT entry_date, shift, SUM(count) as total FROM manpower_detail GROUP BY entry_date, shift"
    ).fetchall()
    c.close()
    total_days = len({r['entry_date'] for r in rows})
    total_mh   = sum(r['total'] * SHIFT_HOURS.get(r['shift'], 0) for r in rows)
    avg = total_mh / total_days if total_days else 0
    return {'total_manhours': total_mh, 'total_days': total_days, 'avg_per_day': avg}


# ── Drawings (stored as BYTEA in database) ─────────────────────────────────────

def save_drawing(title, assembly_mark, original_name, file_bytes, uploaded_by='', rev_no='', date_received=''):
    import uuid
    ext      = original_name.rsplit('.', 1)[-1].lower() if '.' in original_name else 'bin'
    filename = f"{uuid.uuid4().hex}.{ext}"
    c = _conn()
    c.execute("""
        INSERT INTO drawings (title, original_name, filename, assembly_mark, uploaded_by, file_data, rev_no, date_received)
        VALUES (?,?,?,?,?,?,?,?)
    """, (title.strip(), original_name, filename, assembly_mark or '', uploaded_by,
          psycopg2.Binary(file_bytes), rev_no or '', date_received or ''))
    c.commit()
    c.close()


def get_drawings(assembly_mark=None):
    """Return drawing metadata only — file_data excluded to keep list fast."""
    _cols = "id, title, original_name, filename, assembly_mark, uploaded_by, created_at, rev_no, date_received"
    c = _conn()
    if assembly_mark:
        rows = c.execute(
            f"SELECT {_cols} FROM drawings WHERE assembly_mark=? ORDER BY created_at DESC",
            (assembly_mark,)
        ).fetchall()
    else:
        rows = c.execute(
            f"SELECT {_cols} FROM drawings ORDER BY created_at DESC"
        ).fetchall()
    c.close()
    return [dict(r) for r in rows]


def get_drawing_file(did):
    """Fetch file_data bytes for a single drawing — called on demand only."""
    c = _conn()
    row = c.execute("SELECT file_data FROM drawings WHERE id=?", (did,)).fetchone()
    c.close()
    return bytes(row['file_data']) if row and row.get('file_data') else None


def delete_drawing(did):
    c = _conn()
    c.execute("DELETE FROM drawings WHERE id=?", (did,))
    c.commit()
    c.close()
