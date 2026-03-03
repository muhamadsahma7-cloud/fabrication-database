import streamlit as st
import db
import pandas as pd
import base64
from datetime import date, timedelta
from io import BytesIO

APP_VERSION = 'v1.0.0'

st.set_page_config(
    page_title=f'Fabrication Tracker {APP_VERSION}',
    page_icon='🏗️',
    layout='wide',
    initial_sidebar_state='expanded',
)

STAGE_COLOR = {
    'FIT UP':              '#2563eb',
    'WELDING':             '#dc2626',
    'BLASTING & PAINTING': '#7c3aed',
    'SEND TO SITE':        '#ea7c00',
}
STAGE_BADGE = {
    'FIT UP':              '🔵',
    'WELDING':             '🔴',
    'BLASTING & PAINTING': '🟣',
    'SEND TO SITE':        '🟠',
}

@st.cache_data(ttl=120, show_spinner=False)
def _get_marks():
    """Assembly mark list cached for 2 min — avoids a DB round-trip on every rerun."""
    return db.get_marks()

# ── Global CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stSidebar"] {
    background: #1e293b;
    min-width: 280px !important;
    max-width: 280px !important;
}
[data-testid="stSidebar"] * { color: #f1f5f9 !important; }
[data-testid="stSidebar"] .stRadio label {
    font-size: 16px;
    padding: 6px 4px;
}
[data-testid="stSidebar"] h3 { font-size: 18px !important; }
div[data-testid="metric-container"] {
    background: white; border-radius: 8px;
    padding: 12px 16px; box-shadow: 0 1px 3px rgba(0,0,0,.1);
}
.block-container { padding-top: 1.5rem; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# Login
# ══════════════════════════════════════════════════════════════════════════════
def show_login():
    project_name = st.session_state.get('project_name', 'Fabrication Tracker')
    _, mid, _ = st.columns([1, 1.2, 1])
    with mid:
        st.markdown('<br>', unsafe_allow_html=True)
        st.markdown(f'## 🏗️ {project_name}')
        st.caption(APP_VERSION)
        st.divider()
        with st.form('login_form'):
            username = st.text_input('Username', placeholder='Enter username')
            password = st.text_input('Password', type='password', placeholder='Enter password')
            submitted = st.form_submit_button('Login', use_container_width=True, type='primary')
            if submitted:
                if not username or not password:
                    st.error('Enter username and password.')
                else:
                    user = db.authenticate(username, password)
                    if user:
                        st.session_state.user       = user
                        st.session_state.session_id = db.create_session(user['username'], user['role'])
                        default = '📅 Report' if user['role'] == 'viewer' else '✏️ Daily Entry'
                        st.session_state.page  = default
                        st.session_state.queue = []
                        st.rerun()
                    else:
                        st.error('Invalid username or password.')
        st.caption('Default login: admin / admin123')


# ══════════════════════════════════════════════════════════════════════════════
# Sidebar
# ══════════════════════════════════════════════════════════════════════════════
def show_sidebar():
    user = st.session_state.user
    project_name = st.session_state.get('project_name', 'Fabrication Tracker')
    with st.sidebar:
        st.markdown(f'### 🏗️ {project_name}')
        st.caption(APP_VERSION)
        st.divider()
        role_labels = {'admin': 'Administrator', 'user': 'User', 'viewer': 'Viewer (QC)'}
        role_label  = role_labels.get(user['role'], user['role'].capitalize())
        st.markdown(f"👤 **{user['username']}**  ({role_label})")
        st.divider()

        if user['role'] == 'viewer':
            pages = ['📅 Report', '📊 Progress', '🚚 Delivery', '📦 Raw Material', '🖼️ Drawing']
        elif user['role'] == 'admin':
            pages = ['✏️ Daily Entry', '📅 Report', '📊 Progress', '🚚 Delivery',
                     '📦 Raw Material', '🖼️ Drawing', '⚙️ Manage']
        else:
            pages = ['✏️ Daily Entry', '📅 Report', '📊 Progress', '🚚 Delivery',
                     '📦 Raw Material', '🖼️ Drawing']

        default_page = pages[0]
        current = st.session_state.get('page', default_page)
        page = st.radio('Navigation', pages, label_visibility='collapsed',
                        index=pages.index(current) if current in pages else 0)
        st.session_state.page = page

        st.divider()
        st.caption(f'📅 {date.today().strftime("%d %B %Y")}')
        if st.button('🔓 Logout', use_container_width=True):
            if 'session_id' in st.session_state:
                db.end_session(st.session_state.session_id)
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Page: Daily Entry
# ══════════════════════════════════════════════════════════════════════════════
def page_daily_entry():
    st.header('✏️ Daily Entry')

    if 'queue' not in st.session_state:
        st.session_state.queue = []

    marks = _get_marks()

    col_form, col_right = st.columns([1, 1.6])

    # ── Left: Entry Form ──────────────────────────────────────────────────────
    with col_form:
        with st.container(border=True):
            st.subheader('Add Entry')
            entry_date = st.date_input('Date', value=date.today(), key='entry_date')
            mark = st.selectbox('Assembly Mark', [''] + marks, key='entry_mark')

            subs = db.get_sub_assemblies(mark) if mark else []
            subs_selected = st.multiselect(
                'Sub-Assembly Mark', subs, key='entry_sub',
                placeholder='Select one or more (empty = whole assembly)')

            st.write('**Stage**')
            stage_cols = st.columns(2)
            stages_r1  = db.STAGES[:2]
            stages_r2  = db.STAGES[2:]
            if 'sel_stage' not in st.session_state:
                st.session_state.sel_stage = db.STAGES[0]

            for i, s in enumerate(stages_r1):
                if stage_cols[i].button(
                    f'{STAGE_BADGE[s]} {s}',
                    use_container_width=True,
                    type='primary' if st.session_state.sel_stage == s else 'secondary'
                ):
                    st.session_state.sel_stage = s
                    st.rerun()
            for i, s in enumerate(stages_r2):
                if stage_cols[i].button(
                    f'{STAGE_BADGE[s]} {s}',
                    use_container_width=True,
                    type='primary' if st.session_state.sel_stage == s else 'secondary'
                ):
                    st.session_state.sel_stage = s
                    st.rerun()
            stage = st.session_state.sel_stage

            # Auto weight per sub-assembly from parts table
            all_parts = db.get_parts(mark) if mark else []
            def _sub_weight(s):
                pts = [p for p in all_parts if p['sub_assembly_mark'] == s] if s else all_parts
                return round(sum(p['total_weight_kg'] for p in pts), 2)

            if len(subs_selected) >= 2:
                # Multiple subs — auto-calc each, no single override field
                weights_map = {s: _sub_weight(s) for s in subs_selected}
                total_w = sum(weights_map.values())
                st.caption(f"Auto weight: **{total_w:,.2f} kg** total "
                           f"across {len(subs_selected)} sub-assemblies")
            else:
                # 0 or 1 sub — show editable weight field
                s0 = subs_selected[0] if subs_selected else ''
                weight_val = _sub_weight(s0) if mark else 0.0
                weight = st.number_input('Weight (kg)', value=weight_val, min_value=0.0, format='%.2f')
                weights_map = {s0: weight}

            qty     = st.number_input('Qty', value=1, min_value=0, step=1)
            do_no   = ''
            if stage in ('BLASTING & PAINTING', 'SEND TO SITE'):
                do_no = st.text_input('D.O. Number *', placeholder='Required for this stage')
            remarks = st.text_area('Remarks', height=70)

            # Live duplicate warning — check every selected sub
            if mark:
                check_subs  = subs_selected if subs_selected else ['']
                warn_done   = [s for s in check_subs
                               if stage in db.get_completed_stages(mark, s)]
                warn_queued = [s for s in check_subs
                               if stage in {e['stage'] for e in st.session_state.queue
                                            if e['mark'] == mark.upper() and e['sub'] == s.upper()}]
                if warn_done:
                    st.warning(f'⚠️ [{stage}] already recorded for: '
                               f'{", ".join(s or mark for s in warn_done)}')
                elif warn_queued:
                    st.warning(f'⚠️ [{stage}] already in queue for: '
                               f'{", ".join(s or mark for s in warn_queued)}')

            if st.button('➕ Add to Queue', type='primary', use_container_width=True):
                errors = []
                if not mark:
                    errors.append('Assembly Mark is required.')
                if stage in ('BLASTING & PAINTING', 'SEND TO SITE') and not do_no.strip():
                    errors.append(f'D.O. Number is required for [{stage}].')

                check_subs = subs_selected if subs_selected else ['']

                if mark and not errors:
                    stage_idx = db.STAGES.index(stage)
                    for s in check_subs:
                        completed     = db.get_completed_stages(mark, s)
                        queued_stages = {e['stage'] for e in st.session_state.queue
                                         if e['mark'] == mark.upper() and e['sub'] == s.upper()}
                        label = s if s else mark
                        if stage in completed:
                            errors.append(f'[{stage}] already recorded for {label}.')
                        elif stage in queued_stages:
                            errors.append(f'[{stage}] already in queue for {label}.')
                        elif stage_idx > 0:
                            prev_stage = db.STAGES[stage_idx - 1]
                            if prev_stage not in completed and prev_stage not in queued_stages:
                                errors.append(f'{label}: "{prev_stage}" must be completed first.')

                if errors:
                    for e in errors:
                        st.error(e)
                else:
                    for s in check_subs:
                        st.session_state.queue.append({
                            'date':    str(entry_date),
                            'mark':    mark.upper(),
                            'sub':     s.upper(),
                            'stage':   stage,
                            'weight':  weights_map.get(s, 0.0),
                            'qty':     int(qty),
                            'do_no':   do_no.strip(),
                            'remarks': remarks.strip(),
                        })
                    n = len(check_subs)
                    st.success(f'Added {n} item{"s" if n > 1 else ""} to queue.')
                    st.rerun()

    # ── Right: Queue + Saved ──────────────────────────────────────────────────
    with col_right:
        with st.container(border=True):
            st.subheader(f'Queue  ({len(st.session_state.queue)} items)')
            if st.session_state.queue:
                df_q = pd.DataFrame(st.session_state.queue)
                df_q.columns = ['Date', 'Assembly', 'Sub-Assembly', 'Stage',
                                 'Weight (kg)', 'Qty', 'D.O. No.', 'Remarks']
                st.dataframe(df_q, use_container_width=True, hide_index=True)

                c1, c2 = st.columns(2)
                with c1:
                    if st.button('💾 Save All', type='primary', use_container_width=True):
                        for e in st.session_state.queue:
                            db.add_progress(e['date'], e['mark'], e['sub'], e['stage'],
                                            e['weight'], e['qty'], e['remarks'], e['do_no'])
                        count = len(st.session_state.queue)
                        st.session_state.queue = []
                        st.success(f'Saved {count} entries.')
                        st.rerun()
                with c2:
                    if st.button('🗑 Clear Queue', use_container_width=True):
                        st.session_state.queue = []
                        st.rerun()
            else:
                st.info('Queue is empty. Add entries from the form.')

        st.markdown('---')

        with st.container(border=True):
            st.subheader("Today's Saved Entries")
            today_rows = db.search_progress(start=str(date.today()), end=str(date.today()))
            if today_rows:
                df_t = pd.DataFrame(today_rows)[
                    ['id', 'entry_date', 'assembly_mark', 'sub_assembly_mark',
                     'stage', 'delivery_order_no', 'weight_kg', 'qty', 'remarks']]
                df_t.columns = ['ID', 'Date', 'Assembly', 'Sub-Assembly',
                                 'Stage', 'D.O. No.', 'Weight (kg)', 'Qty', 'Remarks']
                st.dataframe(df_t, use_container_width=True, hide_index=True)

                with st.form('del_today'):
                    del_id = st.number_input('Delete entry by ID', min_value=0, step=1, value=0)
                    if st.form_submit_button('🗑 Delete', type='secondary'):
                        if del_id > 0:
                            db.delete_progress(int(del_id))
                            st.success(f'Deleted entry #{int(del_id)}')
                            st.rerun()
            else:
                st.info('No entries saved today.')

    st.markdown('---')
    with st.container(border=True):
        st.subheader('👷 Daily Manpower')
        mp_date  = st.date_input('Date', value=date.today(), key='mp_date')
        existing = db.get_manpower_grid(mp_date)

        # Header row
        hcols = st.columns([1.6, 1, 1, 1, 1, 1])
        hcols[0].markdown('**Worker Type**')
        for i, label in enumerate(db.SHIFT_LABELS):
            hcols[i + 1].markdown(f'**{label}**')

        # Grid: one row per worker type, one column per shift
        new_grid = {}
        day_mh = 0
        total_workers = 0
        for wtype in db.WORKER_TYPES:
            wslug = wtype.lower().replace(' ', '_')
            row   = st.columns([1.6, 1, 1, 1, 1, 1])
            row[0].markdown(wtype)
            new_grid[wtype] = {}
            for j, shift_key in enumerate(db.SHIFT_KEYS):
                val = existing.get(wtype, {}).get(shift_key, 0)
                count = row[j + 1].number_input(
                    '', min_value=0, value=val, step=1,
                    key=f'mp_{wslug}_{shift_key}', label_visibility='collapsed'
                )
                new_grid[wtype][shift_key] = count
                day_mh       += count * db.SHIFT_HOURS[shift_key]
                total_workers += count

        st.caption(f"Manhours for this day: **{day_mh:,.1f} hrs** · "
                   f"Total workers: **{total_workers}**")

        if st.button('💾 Save Manpower', type='primary', use_container_width=True, key='mp_save'):
            db.save_manpower_grid(mp_date, new_grid)
            st.success(f'Manpower saved for {mp_date}')
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Page: Report
# ══════════════════════════════════════════════════════════════════════════════
def page_report():
    st.header('📅 Report')

    # ── Today's Activity ──────────────────────────────────────────────────────
    today_rows = db.search_progress(start=str(date.today()), end=str(date.today()))
    today_by_stage = {s: sum(r['weight_kg'] for r in today_rows if r['stage'] == s)
                      for s in db.STAGES}

    # Single aggregate query replaces loading all FIT UP / WELDING rows
    stage_stats  = db.get_stage_daily_stats()
    fitup_stats  = stage_stats.get('FIT UP',  {'total_kg': 0, 'days': 0, 'avg_per_day': 0})
    weld_stats   = stage_stats.get('WELDING', {'total_kg': 0, 'days': 0, 'avg_per_day': 0})

    st.markdown(f"**Today's Activity** — {date.today().strftime('%d %b %Y')}")
    tcols = st.columns(len(db.STAGES))
    for i, s in enumerate(db.STAGES):
        with tcols[i]:
            st.metric(f'{STAGE_BADGE[s]} {s}', f'{today_by_stage[s]:,.1f} kg')
            if s == 'FIT UP':
                d = fitup_stats['days']
                st.metric('Avg/Day', f'{fitup_stats["avg_per_day"]:,.1f} kg',
                          f'{fitup_stats["total_kg"]:,.1f} kg ÷ {d} day{"s" if d!=1 else ""}')
            elif s == 'WELDING':
                d = weld_stats['days']
                st.metric('Avg/Day', f'{weld_stats["avg_per_day"]:,.1f} kg',
                          f'{weld_stats["total_kg"]:,.1f} kg ÷ {d} day{"s" if d!=1 else ""}')

    mh = db.get_manhour_summary()
    today_grid = db.get_manpower_grid(date.today())
    today_mh   = sum(
        count * db.SHIFT_HOURS[sk]
        for shifts in today_grid.values()
        for sk, count in shifts.items()
    )
    mh_cols = st.columns(3)
    with mh_cols[0]:
        st.metric("Today's Manhours", f"{today_mh:,.1f} hrs")
    with mh_cols[1]:
        st.metric('Avg Manhour/Day',
                  f"{mh['avg_per_day']:,.1f} hrs",
                  f"{mh['total_manhours']:,.1f} hrs ÷ {mh['total_days']} day{'s' if mh['total_days']!=1 else ''}"
                  if mh['total_days'] else None)
    with mh_cols[2]:
        st.metric('Total Manhours', f"{mh['total_manhours']:,.1f} hrs")
    st.divider()

    # Placeholder: summary metrics will be injected here (above filters)
    summary_container = st.container()

    # ── Filters — always visible, directly above table ────────────────────────
    c1, c2, c3, c4 = st.columns([1, 1, 1.2, 1.2])
    with c1:
        start = st.date_input('From', value=date.today() - timedelta(days=7), key='rpt_start')
    with c2:
        end = st.date_input('To', value=date.today(), key='rpt_end')
    with c3:
        asm_filter = st.selectbox('Assembly', ['All'] + _get_marks(), key='rpt_asm')
    with c4:
        stage_filter = st.selectbox('Stage', ['All'] + db.STAGES, key='rpt_stage')

    bc1, bc2 = st.columns([1, 1])
    with bc1:
        load = st.button('🔍 Load by Date', type='primary', use_container_width=True)
    with bc2:
        load_all = st.button('📋 Show All', use_container_width=True)

    if 'report_rows' not in st.session_state:
        st.session_state.report_rows = []

    asm = None if asm_filter == 'All' else asm_filter
    stg = None if stage_filter == 'All' else stage_filter

    if load:
        st.session_state.report_rows = db.search_progress(
            stage=stg, assembly_mark=asm, start=str(start), end=str(end))
    if load_all:
        st.session_state.report_rows = db.search_progress(stage=stg, assembly_mark=asm)

    rows = st.session_state.report_rows
    if rows:
        summary       = db.get_summary()
        project_total = summary.get('total', 0) or 0
        stage_totals  = {s: sum(r['weight_kg'] for r in rows if r['stage'] == s)
                         for s in db.STAGES}

        # Fill the placeholder above the filters
        with summary_container:
            metric_cols = st.columns(len(db.STAGES) + 1)
            with metric_cols[0]:
                st.metric('Total Weight (Project)', f'{project_total:,.1f} kg')
            for i, s in enumerate(db.STAGES):
                pct = min(stage_totals[s] / project_total * 100, 100) if project_total else 0
                with metric_cols[i + 1]:
                    st.metric(f'{STAGE_BADGE[s]} {s}',
                              f'{stage_totals[s]:,.1f} kg', f'{pct:.1f}%')
            st.divider()

        df = pd.DataFrame(rows)[
            ['entry_date', 'assembly_mark', 'sub_assembly_mark', 'stage',
             'delivery_order_no', 'weight_kg', 'qty', 'remarks']]
        df.columns = ['Date', 'Assembly', 'Sub-Assembly', 'Stage',
                      'D.O. No.', 'Weight (kg)', 'Qty', 'Remarks']
        st.dataframe(df, use_container_width=True, hide_index=True)

        ec1, ec2 = st.columns(2)
        with ec1:
            st.download_button('📥 Export CSV',
                               df.to_csv(index=False).encode('utf-8'),
                               f'report_{start}_{end}.csv', 'text/csv',
                               use_container_width=True)
        with ec2:
            buf = BytesIO()
            df.to_excel(buf, index=False, engine='openpyxl')
            st.download_button('📥 Export Excel', buf.getvalue(),
                               f'report_{start}_{end}.xlsx',
                               'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               use_container_width=True)
    else:
        st.info('Click **Load by Date** or **Show All** to display data.')


# ══════════════════════════════════════════════════════════════════════════════
# Page: Progress
# ══════════════════════════════════════════════════════════════════════════════
def page_progress():
    st.header('📊 Progress Overview')

    summary       = db.get_summary()
    project_total = summary.get('total', 0) or 0

    met_cols = st.columns(len(db.STAGES))
    for i, s in enumerate(db.STAGES):
        done = summary.get(s, 0) or 0
        pct  = min(done / project_total * 100, 100) if project_total else 0
        with met_cols[i]:
            st.metric(f'{STAGE_BADGE[s]} {s}', f'{done:,.1f} kg', f'{pct:.1f}%')

    st.divider()

    rows = db.get_cumulative_by_sub()
    if not rows:
        st.info('No progress data yet.')
        return

    df = pd.DataFrame(rows).rename(columns={
        'assembly_mark':    'Assembly',
        'sub_assembly_mark':'Sub-Assembly',
        'total_weight_kg':  'Total (kg)',
        'fitup':            'Fit Up (kg)',
        'welding':          'Welding (kg)',
        'blasting':         'Blast/Paint (kg)',
        'sendsite':         'Send to Site (kg)',
    })

    def pct_str(row, col):
        t = row['Total (kg)']
        return f"{min(row[col]/t*100, 100):.1f}%" if t else '—'

    df['Fit Up %']       = df.apply(lambda r: pct_str(r, 'Fit Up (kg)'),       axis=1)
    df['Welding %']      = df.apply(lambda r: pct_str(r, 'Welding (kg)'),      axis=1)
    df['Blast/Paint %']  = df.apply(lambda r: pct_str(r, 'Blast/Paint (kg)'),  axis=1)
    df['Send to Site %'] = df.apply(lambda r: pct_str(r, 'Send to Site (kg)'), axis=1)

    display_cols = [
        'Assembly', 'Sub-Assembly', 'Total (kg)',
        'Fit Up (kg)', 'Fit Up %',
        'Welding (kg)', 'Welding %',
        'Blast/Paint (kg)', 'Blast/Paint %',
        'Send to Site (kg)', 'Send to Site %',
    ]
    st.dataframe(df[display_cols], use_container_width=True, hide_index=True)

    # ── Download Excel ─────────────────────────────────────────────────────────
    import openpyxl as _xl
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side

    prog_wb  = _xl.Workbook()
    prog_ws  = prog_wb.active
    prog_ws.title = 'Progress'

    hdr_cols = [
        ('Assembly Mark',        18, None),
        ('Sub-Assembly Mark',    22, None),
        ('Total Weight (kg)',    16, '#,##0.00'),
        ('Fit Up (kg)',          14, '#,##0.00'),
        ('Fit Up %',             10, '0.0%'),
        ('Welding (kg)',         14, '#,##0.00'),
        ('Welding %',            10, '0.0%'),
        ('Blast/Paint (kg)',     16, '#,##0.00'),
        ('Blast/Paint %',        12, '0.0%'),
        ('Send to Site (kg)',    16, '#,##0.00'),
        ('Send to Site %',       14, '0.0%'),
    ]

    hdr_fill = _Fill('solid', fgColor='1E3A5F')
    hdr_font = _Font(bold=True, color='FFFFFF', size=11)
    hdr_aln  = _Align(horizontal='center', vertical='center')
    num_aln  = _Align(horizontal='right',  vertical='center')
    txt_aln  = _Align(horizontal='left',   vertical='center')
    thin     = _Side(style='thin', color='CCCCCC')
    border   = _Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for col_idx, (col_name, col_w, _) in enumerate(hdr_cols, 1):
        cell = prog_ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_aln
        cell.border    = border
        prog_ws.column_dimensions[cell.column_letter].width = col_w

    prog_ws.row_dimensions[1].height = 20

    def _safe_pct(part, total):
        try:
            return min(float(part) / float(total), 1.0) if total else 0.0
        except Exception:
            return 0.0

    # Data rows
    stage_key = {'Fit Up': 'fitup', 'Welding': 'welding',
                 'Blast/Paint': 'blasting', 'Send to Site': 'sendsite'}
    for r_idx, r in enumerate(rows, 2):
        total = r.get('total_weight_kg') or 0
        fitup    = r.get('fitup')    or 0
        welding  = r.get('welding')  or 0
        blasting = r.get('blasting') or 0
        sendsite = r.get('sendsite') or 0

        row_vals = [
            r.get('assembly_mark',     ''),
            r.get('sub_assembly_mark', ''),
            total,
            fitup,    _safe_pct(fitup,    total),
            welding,  _safe_pct(welding,  total),
            blasting, _safe_pct(blasting, total),
            sendsite, _safe_pct(sendsite, total),
        ]
        row_fill = _Fill('solid', fgColor='F0F4FA') if r_idx % 2 == 0 else _Fill('solid', fgColor='FFFFFF')
        for col_idx, (val, (_, _, num_fmt)) in enumerate(zip(row_vals, hdr_cols), 1):
            cell = prog_ws.cell(row=r_idx, column=col_idx, value=val)
            cell.border    = border
            cell.fill      = row_fill
            if num_fmt:
                cell.number_format = num_fmt
                cell.alignment     = num_aln
            else:
                cell.alignment = txt_aln

    prog_ws.freeze_panes = 'A2'
    prog_ws.auto_filter.ref = prog_ws.dimensions

    prog_buf = BytesIO()
    prog_wb.save(prog_buf)

    st.download_button(
        '📥 Download Excel',
        prog_buf.getvalue(),
        f'progress_{date.today()}.xlsx',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        use_container_width=True,
        type='primary',
    )


# ══════════════════════════════════════════════════════════════════════════════
# Page: Delivery
# ══════════════════════════════════════════════════════════════════════════════
def page_delivery():
    st.header('🚚 Delivery Log')

    c1, c2 = st.columns([1, 1])
    with c1:
        start = st.date_input('From', value=date.today() - timedelta(days=30), key='del_start')
    with c2:
        end = st.date_input('To', value=date.today(), key='del_end')

    all_rows = db.get_deliveries()
    rows = [r for r in all_rows
            if str(start) <= r['entry_date'] <= str(end)] if all_rows else []

    if rows:
        df = pd.DataFrame(rows)[
            ['entry_date', 'assembly_mark', 'sub_assembly_mark', 'stage',
             'delivery_order_no', 'weight_kg', 'qty', 'remarks']]
        df.columns = ['Date', 'Assembly', 'Sub-Assembly', 'Type',
                      'D.O. No.', 'Weight (kg)', 'Qty', 'Remarks']
        st.dataframe(df, use_container_width=True, hide_index=True)

        ec1, ec2 = st.columns(2)
        with ec1:
            st.download_button('📥 Export CSV',
                               df.to_csv(index=False).encode('utf-8'),
                               'delivery_log.csv', 'text/csv',
                               use_container_width=True)
        with ec2:
            buf = BytesIO()
            df.to_excel(buf, index=False, engine='openpyxl')
            st.download_button('📥 Export Excel', buf.getvalue(),
                               'delivery_log.xlsx',
                               'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               use_container_width=True)
    else:
        st.info('No delivery records in this date range.')


# ══════════════════════════════════════════════════════════════════════════════
# Page: Manage (admin only)
# ══════════════════════════════════════════════════════════════════════════════
def page_manage():
    st.header('⚙️ Manage Data')

    tab_import, tab_export, tab_users, tab_online, tab_settings, tab_danger = st.tabs(
        ['📥 Import Excel', '📤 Export Master', '👥 Users', '🟢 Online', '⚙️ Settings', '⚠️ Danger Zone'])

    # ── Import ────────────────────────────────────────────────────────────────
    with tab_import:
        st.subheader('Import Master Database from Excel')

        # ── Template download ──────────────────────────────────────────────────
        tpl_buf = BytesIO()
        import openpyxl as _xl
        from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align
        tpl_wb = _xl.Workbook()
        tpl_ws = tpl_wb.active
        tpl_ws.title = 'Master Database'
        tpl_headers = ['Assembly Mark', 'Sub Assembly', 'Part Mark', 'No.',
                       'Name', 'Profile', 'kg/m', 'Length (mm)', 'Weight (kg)',
                       'Profile 2', 'Grade', 'Remark',
                       'FIT UP (kg)', 'FIT UP Date',
                       'WELDING (kg)', 'WELDING Date',
                       'BLASTING & PAINTING (kg)', 'BLASTING & PAINTING Date',
                       'SEND TO SITE (kg)', 'SEND TO SITE Date']
        hdr_fill = _Fill('solid', fgColor='1E3A5F')
        hdr_font = _Font(bold=True, color='FFFFFF')
        for col, h in enumerate(tpl_headers, 1):
            cell = tpl_ws.cell(row=1, column=col, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = _Align(horizontal='center')
            tpl_ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 14)
        tpl_wb.save(tpl_buf)
        st.download_button(
            '📄 Download Template (.xlsx)',
            tpl_buf.getvalue(),
            'master_database_template.xlsx',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        st.divider()
        st.info(
            '**Upload your revised Master Database Excel file.**  \n'
            'This will overwrite all existing assemblies & parts with the new data.  \n'
            '✅ All daily progress records are kept safe.'
        )
        uploaded = st.file_uploader('Choose Excel file (.xlsx)', type=['xlsx', 'xls'])
        if st.button('📥 Import & Overwrite', type='primary', use_container_width=True,
                     disabled=uploaded is None):
            file_bytes = uploaded.read()
            part_count, prog_count, err = db.replace_import_excel(file_bytes)
            if err:
                st.error(f'Import failed: {err}')
            else:
                st.success(f'✅ Imported {part_count} parts and {prog_count} progress records.')


    # ── Export ────────────────────────────────────────────────────────────────
    with tab_export:
        st.subheader('Export Master Database')
        st.caption('Parts list with cumulative progress per stage.')
        rows = db.get_master_export()
        if rows:
            df = pd.DataFrame(rows)
            ec1, ec2 = st.columns(2)
            with ec1:
                st.download_button('📥 Download CSV',
                                   df.to_csv(index=False).encode('utf-8'),
                                   f'master_database_{date.today()}.csv', 'text/csv',
                                   use_container_width=True, type='primary')
            with ec2:
                buf = BytesIO()
                df.to_excel(buf, index=False, engine='openpyxl')
                st.download_button('📥 Download Excel', buf.getvalue(),
                                   f'master_database_{date.today()}.xlsx',
                                   'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                   use_container_width=True, type='primary')
            st.dataframe(df, use_container_width=True, hide_index=True)
        else:
            st.info('No parts data to export.')

    # ── Users ─────────────────────────────────────────────────────────────────
    with tab_users:
        st.subheader('User List')
        users = db.get_users()
        if users:
            df_u = pd.DataFrame(users)
            df_u['active'] = df_u['active'].map({1: '✅ Active', 0: '❌ Inactive'})
            df_u.columns   = ['ID', 'Username', 'Role', 'Status']
            st.dataframe(df_u, use_container_width=True, hide_index=True)
        else:
            st.info('No users found.')

        st.divider()
        uc1, uc2 = st.columns(2)

        with uc1:
            with st.form('add_user'):
                st.subheader('Add User')
                uname = st.text_input('Username')
                pwd   = st.text_input('Password', type='password')
                role  = st.selectbox('Role', ['user', 'admin', 'viewer'])
                if st.form_submit_button('Add User', type='primary'):
                    if not uname or not pwd:
                        st.error('Username and password required.')
                    elif db.add_user(uname, pwd, role):
                        st.success(f'User "{uname}" added.')
                        st.rerun()
                    else:
                        st.error('Username already exists.')

        with uc2:
            with st.form('reset_pwd'):
                st.subheader('Reset Password')
                uid     = st.number_input('User ID', min_value=1, step=1, value=1)
                new_pwd = st.text_input('New Password', type='password')
                if st.form_submit_button('Reset Password'):
                    if new_pwd:
                        db.update_user_password(int(uid), new_pwd)
                        st.success('Password updated.')
                    else:
                        st.error('Password cannot be empty.')

        st.divider()
        with st.form('toggle_user'):
            st.subheader('Toggle Active / Delete')
            act_uid = st.number_input('User ID', min_value=1, step=1, value=1, key='act_uid')
            tc1, tc2 = st.columns(2)
            with tc1:
                toggle = st.form_submit_button('🔄 Toggle Active')
            with tc2:
                delete = st.form_submit_button('🗑 Delete User')
            if toggle:
                db.toggle_user_active(int(act_uid))
                st.success(f'Toggled user {int(act_uid)}.')
                st.rerun()
            if delete:
                db.delete_user_entry(int(act_uid))
                st.success(f'Deleted user {int(act_uid)}.')
                st.rerun()

    # ── Online / Session Tracking ─────────────────────────────────────────────
    with tab_online:
        st.subheader('🟢 Currently Online')
        st.caption('Users active in the last 10 minutes (GMT+8).')
        active = db.get_active_sessions(minutes=10)
        if active:
            role_labels = {'admin': '🔴 Admin', 'user': '🟡 User', 'viewer': '🔵 Viewer'}
            for s in active:
                rl  = role_labels.get(s['role'], s['role'].capitalize())
                col1, col2, col3 = st.columns([1.2, 1.5, 1.5])
                col1.markdown(f"**{s['username']}**")
                col2.caption(f"{rl}")
                col3.caption(f"Last seen: {s['last_seen']} GMT+8")
        else:
            st.info('No active sessions in the last 10 minutes.')

        st.divider()
        st.subheader('📋 Login History')
        history = db.get_login_history(limit=100)
        if history:
            df_h = pd.DataFrame(history)
            df_h['active'] = df_h['active'].map({1: '🟢 Online', 0: '⚫ Logged out'})
            df_h.columns   = ['Username', 'Role', 'Login Time (GMT+8)', 'Last Seen (GMT+8)', 'Status']
            st.dataframe(df_h, use_container_width=True, hide_index=True)
        else:
            st.info('No login history yet.')

    # ── Settings ──────────────────────────────────────────────────────────────
    with tab_settings:
        st.subheader('Project Settings')
        current_name = st.session_state.get('project_name', 'Fabrication Tracker')
        with st.form('project_name_form'):
            new_name = st.text_input('Project Name', value=current_name,
                                     placeholder='e.g. Ulu Tiram Station Fabrication')
            if st.form_submit_button('💾 Save Project Name', type='primary'):
                if new_name.strip():
                    db.set_project_name(new_name)
                    st.session_state.project_name = new_name.strip()
                    st.success(f'Project name updated to: **{new_name.strip()}**')
                    st.rerun()
                else:
                    st.error('Project name cannot be empty.')

    # ── Danger Zone ───────────────────────────────────────────────────────────
    with tab_danger:
        st.subheader('⚠️ Danger Zone')
        st.error('**Clear All Database** will permanently delete ALL assemblies, '
                 'parts, and progress records. This cannot be undone.')
        confirm = st.checkbox('I understand this will delete everything permanently.')
        if confirm:
            if st.button('🗑 Clear All Database', type='primary'):
                db.clear_all_data()
                st.session_state.report_rows = []
                st.success('All database records have been cleared.')
                st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════
# Page: Raw Material Delivery
# ══════════════════════════════════════════════════════════════════════════════
def page_raw_material():
    st.header('📦 Raw Material Delivery')
    role = st.session_state.user['role']

    if role != 'viewer':
        col_add, col_imp = st.columns([1.4, 1])

        with col_add:
            with st.container(border=True):
                st.subheader('Add Received Material')
                with st.form('raw_form', clear_on_submit=True):
                    c1, c2 = st.columns(2)
                    with c1:
                        recv_date   = st.date_input('Received Date', value=date.today())
                        do_no       = st.text_input('D.O. Number', placeholder='Delivery Order No.')
                        description = st.text_input('Description', placeholder='e.g. UB 356x171x45')
                    with c2:
                        grade    = st.text_input('Material Grade', placeholder='e.g. S275, S355')
                        qty      = st.number_input('Qty', min_value=0.0, format='%.2f')
                        total_kg = st.number_input('Total kg', min_value=0.0, format='%.2f')
                        remark   = st.text_area('Remark', height=70)
                    if st.form_submit_button('➕ Add', type='primary', use_container_width=True):
                        if not description:
                            st.error('Description is required.')
                        else:
                            db.add_raw_material(recv_date, do_no, description, grade, qty, total_kg, remark)
                            st.success('Raw material entry added.')
                            st.rerun()

        with col_imp:
            with st.container(border=True):
                st.subheader('Import from Excel')

                # Template download
                rm_tpl_buf = BytesIO()
                import openpyxl as _xl2
                from openpyxl.styles import Font as _Font2, PatternFill as _Fill2, Alignment as _Align2
                rm_tpl_wb = _xl2.Workbook()
                rm_tpl_ws = rm_tpl_wb.active
                rm_tpl_ws.title = 'Raw Material'
                rm_tpl_headers = ['Received Date', 'D.O. Number', 'Description', 'Grade', 'Qty', 'Total kg', 'Remark']
                rm_hdr_fill = _Fill2('solid', fgColor='1E3A5F')
                rm_hdr_font = _Font2(bold=True, color='FFFFFF')
                for col, h in enumerate(rm_tpl_headers, 1):
                    cell = rm_tpl_ws.cell(row=1, column=col, value=h)
                    cell.fill = rm_hdr_fill
                    cell.font = rm_hdr_font
                    cell.alignment = _Align2(horizontal='center')
                    rm_tpl_ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)
                rm_tpl_wb.save(rm_tpl_buf)
                st.download_button(
                    '📄 Download Template (.xlsx)',
                    rm_tpl_buf.getvalue(),
                    'raw_material_template.xlsx',
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True,
                )

                st.caption('Columns: **Received Date · D.O. Number · Description · Grade · Qty · Total kg · Remark**')
                uploaded = st.file_uploader('Choose Excel file (.xlsx)', type=['xlsx', 'xls'],
                                            key='rm_upload')
                if st.button('📥 Import', type='primary', use_container_width=True,
                             disabled=uploaded is None):
                    file_bytes = uploaded.read()
                    count, err = db.import_raw_materials_excel(file_bytes)
                    if err:
                        st.error(f'Import failed: {err}')
                    else:
                        st.success(f'Imported {count} records.')
                        st.session_state.rm_rows = []
                        st.rerun()

    st.markdown('---')

    # ── Overall summary (always visible) ──────────────────────────────────────
    summ = db.get_raw_material_summary()
    sm1, sm2, sm3 = st.columns(3)
    sm1.metric('Total Entries', f"{summ['entries']:,}")
    sm2.metric('Total Qty Received', f"{summ['total_qty']:,.2f}")
    sm3.metric('Total kg Received', f"{summ['total_kg']:,.2f} kg")

    st.divider()

    c1, c2 = st.columns(2)
    with c1:
        start = st.date_input('From', value=date.today() - timedelta(days=30), key='rm_start')
    with c2:
        end = st.date_input('To', value=date.today(), key='rm_end')

    bc1, bc2 = st.columns(2)
    with bc1:
        load     = st.button('🔍 Load by Date', type='primary', use_container_width=True)
    with bc2:
        load_all = st.button('📋 Show All', use_container_width=True)

    if 'rm_rows' not in st.session_state:
        st.session_state.rm_rows = []

    if load:
        st.session_state.rm_rows = db.get_raw_materials(str(start), str(end))
    if load_all:
        st.session_state.rm_rows = db.get_raw_materials()

    rows = st.session_state.rm_rows
    if rows:
        # ── Filtered summary ──────────────────────────────────────────────────
        filt_qty = sum(r.get('qty', 0) for r in rows)
        filt_kg  = sum(r.get('total_kg', 0) for r in rows)
        fc1, fc2, fc3 = st.columns(3)
        fc1.metric('Entries (filtered)', str(len(rows)))
        fc2.metric('Qty (filtered)', f'{filt_qty:,.2f}')
        fc3.metric('Total kg (filtered)', f'{filt_kg:,.2f} kg')

        st.markdown('')
        df = pd.DataFrame(rows)[['id', 'received_date', 'do_no', 'description', 'grade', 'qty', 'total_kg', 'remark']]
        df.columns = ['ID', 'Received Date', 'D.O. No.', 'Description', 'Grade', 'Qty', 'Total kg', 'Remark']
        st.dataframe(df, use_container_width=True, hide_index=True)

        ec1, ec2 = st.columns(2)
        with ec1:
            st.download_button('📥 Export CSV',
                               df.to_csv(index=False).encode('utf-8'),
                               f'raw_material_{date.today()}.csv', 'text/csv',
                               use_container_width=True)
        with ec2:
            buf = BytesIO()
            df.to_excel(buf, index=False, engine='openpyxl')
            st.download_button('📥 Export Excel', buf.getvalue(),
                               f'raw_material_{date.today()}.xlsx',
                               'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               use_container_width=True)

        if role != 'viewer':
            with st.form('del_rm'):
                del_id = st.number_input('Delete entry by ID', min_value=0, step=1, value=0)
                if st.form_submit_button('🗑 Delete', type='secondary'):
                    if del_id > 0:
                        db.delete_raw_material(int(del_id))
                        st.success(f'Deleted entry #{int(del_id)}')
                        st.rerun()
    else:
        st.info('Click **Load by Date** or **Show All** to display records.')


# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# Page: Drawing
# ══════════════════════════════════════════════════════════════════════════════
def page_drawing():
    st.header('🖼️ Drawing')
    role = st.session_state.user['role']

    # ── Upload (non-viewer only) ───────────────────────────────────────────────
    if role != 'viewer':
        with st.container(border=True):
            st.subheader('Upload Drawings')
            uc1, uc2, uc3 = st.columns([1, 1, 1.5])
            with uc1:
                drw_rev  = st.text_input('Rev No.', key='drw_rev', placeholder='e.g. A, B, C1')
            with uc2:
                drw_date = st.date_input('Date Received', value=date.today(), key='drw_date')
            with uc3:
                drw_asm  = st.selectbox('Assembly Mark (optional)',
                                        [''] + _get_marks(), key='drw_asm')

            drw_files = st.file_uploader(
                'Files (PDF, PNG, JPG) — select one or multiple',
                type=['pdf', 'png', 'jpg', 'jpeg'],
                accept_multiple_files=True,
                key='drw_file'
            )

            if drw_files:
                preview = [{'Title (auto)': f.name.rsplit('.', 1)[0], 'File': f.name}
                           for f in drw_files]
                st.caption(f'{len(drw_files)} file(s) selected — titles taken from filenames:')
                st.dataframe(preview, use_container_width=True, hide_index=True)

            if st.button('📤 Upload All', type='primary', use_container_width=True, key='drw_upload'):
                if not drw_files:
                    st.error('Please select at least one file.')
                else:
                    for f in drw_files:
                        title = f.name.rsplit('.', 1)[0]   # filename without extension
                        db.save_drawing(
                            title, drw_asm,
                            f.name, f.read(),
                            st.session_state.user['username'],
                            drw_rev.strip(),
                            str(drw_date),
                        )
                    n = len(drw_files)
                    st.success(f'Uploaded {n} drawing{"s" if n > 1 else ""}.')
                    st.rerun()

    st.divider()

    # ── Filter ────────────────────────────────────────────────────────────────
    filter_asm = st.selectbox('Filter by Assembly', ['All'] + _get_marks(), key='drw_filter')
    asm_f      = None if filter_asm == 'All' else filter_asm
    drawings   = db.get_drawings(assembly_mark=asm_f)

    if not drawings:
        st.info('No drawings uploaded yet.')
        return

    # ── Drawing list ──────────────────────────────────────────────────────────
    for drw in drawings:
        ext   = drw['original_name'].rsplit('.', 1)[-1].lower() if '.' in drw['original_name'] else ''
        label = f"📄 {drw['title']}"
        if drw.get('rev_no'):
            label += f"  ·  Rev {drw['rev_no']}"
        if drw.get('date_received'):
            label += f"  ·  {drw['date_received']}"
        if drw['assembly_mark']:
            label += f"  ·  {drw['assembly_mark']}"

        with st.expander(label):
            meta_parts = []
            if drw.get('rev_no'):
                meta_parts.append(f"**Rev:** {drw['rev_no']}")
            if drw.get('date_received'):
                meta_parts.append(f"**Date Received:** {drw['date_received']}")
            if drw.get('assembly_mark'):
                meta_parts.append(f"**Assembly:** {drw['assembly_mark']}")
            if meta_parts:
                st.caption('  ·  '.join(meta_parts))

            # File data loaded on demand — avoids fetching all BYTEAs on list render
            load_key = f'drw_loaded_{drw["id"]}'
            if not st.session_state.get(load_key):
                if st.button('📂 Load Drawing', key=f'load_{drw["id"]}',
                             use_container_width=True):
                    st.session_state[load_key] = True
                    st.rerun()
            else:
                file_bytes = db.get_drawing_file(drw['id'])
                if not file_bytes:
                    st.warning('File data not found on server.')
                elif ext in ('png', 'jpg', 'jpeg'):
                    st.image(file_bytes, use_container_width=True)
                elif ext == 'pdf':
                    # Chrome blocks data: URIs in iframes — use download button instead
                    st.info('PDF preview is not supported in Chrome. Use the download button below to open the file.')
                if file_bytes:
                    st.download_button(
                        f'📥 Download {drw["original_name"]}',
                        file_bytes, drw['original_name'],
                        key=f'dl_{drw["id"]}',
                        use_container_width=True,
                        type='primary',
                    )

            if role == 'admin':
                if st.button('🗑 Delete', key=f'del_drw_{drw["id"]}', type='secondary'):
                    db.delete_drawing(drw['id'])
                    st.rerun()


@st.cache_resource(show_spinner='Connecting to database…')
def _init_db(_schema_v=4):
    """Run schema init once per server lifecycle. Increment _schema_v to bust cache."""
    db.init()


def main():
    try:
        _init_db()
    except KeyError as e:
        st.error(f"⚠️ Missing Streamlit secret: **{e}**")
        st.info("Go to Streamlit Cloud → your app → ⋮ Settings → **Secrets** and add:\n\n"
                "```toml\n"
                'database_url = "postgresql://postgres.fwynwlagixbisxybdgya:5%405wsHgUiWwgyF5@aws-1-ap-southeast-2.pooler.supabase.com:5432/postgres?sslmode=require"\n'
                "```")
        st.stop()
    except Exception as e:
        st.error("⚠️ Cannot connect to database.")
        st.info("**Possible causes:**\n"
                "- Supabase free-tier project is **paused** — go to [supabase.com](https://supabase.com) "
                "→ your project → click **Restore** to wake it\n"
                "- Wrong credentials in Streamlit secrets\n\n"
                f"Technical detail: `{type(e).__name__}: {e}`")
        st.stop()

    # Cache project name in session state — avoids a DB hit on every rerun
    if 'project_name' not in st.session_state:
        st.session_state.project_name = db.get_project_name()

    if 'user' not in st.session_state:
        show_login()
        return

    # ── Heartbeat: update last_seen every 30 s ────────────────────────────────
    import time as _time
    if 'session_id' in st.session_state:
        now_ts = _time.time()
        if now_ts - st.session_state.get('_last_hb', 0) >= 30:
            db.update_session_heartbeat(st.session_state.session_id)
            st.session_state._last_hb = now_ts

    show_sidebar()

    page = st.session_state.get('page', '✏️ Daily Entry')

    if '✏️' in page:
        page_daily_entry()
    elif '📅' in page:
        page_report()
    elif '📊' in page:
        page_progress()
    elif '🚚' in page:
        page_delivery()
    elif '📦' in page:
        page_raw_material()
    elif '🖼️' in page:
        page_drawing()
    elif '⚙️' in page:
        page_manage()


if __name__ == '__main__':
    main()
